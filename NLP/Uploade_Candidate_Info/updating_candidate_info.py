import openpyxl
import mysql.connector
from mysql.connector import Error
import csv
import logging
import json
from datetime import datetime
import time
import os
import requests


FIELD_NAMES = ['first name', 'last name', 'mob no', 'email', 'created time', 'test score', 'linked resume',
               'notice period', 'current ctc', 'expected ctc', 'skill1', 'skill1 ID', 'skill1 years', 'skill2',
               'skill2 ID', 'skill2 years', 'skill3', 'skill3 ID', 'skill3 years', 'remark']

CONFIG_FILE_PATH = "/home/ubuntu/Python-Projects/NLP/Uploade_Candidate_Info/config.json"
# CONFIG_FILE_PATH = "config.json"

HOST_NAME = ""
USER_NAME = ""
PASSWORD = ""
DB_NAME = ""

RESUME_DIRECTORY_PATH = ""

TABLE_NAME = "upload"
TABLE_COLUMN_ID = "Id"
TABLE_COLUMN_PATH = "Path"
TABLE_COLUMN_PROCESS_STATE = "ProcessedState"
TABLE_COLUMN_JOB_ID = "JobId"

CANDIDATE_TABLE_NAME = "candidate"
CANDIDATE_TABLE_FIRST_NAME = "FirstName"
CANDIDATE_TABLE_LAST_NAME = "LastName"
CANDIDATE_TABLE_DISPLAY_NAME = "DisplayName"
CANDIDATE_TABLE_COLUMN_ID = "Id"
CANDIDATE_TABLE_COLUMN_MOB_NO = "Mobile"
CANDIDATE_TABLE_COLUMN_CCTC = "CTC"
CANDIDATE_TABLE_COLUMN_ECTC = "ECTC"
CANDIDATE_TABLE_COLUMN_EMAIL_ID = "EmailId"
CANDIDATE_TABLE_COLUMN_NOTICE_PERIOD = "NoticePeriod"
CANDIDATE_TABLE_COLUMN_CREATED_TIME = "CreatedTime"
CANDIDATE_TABLE_COLUMN_LAST_UPDATED_TIME = "LastUpdatedTime"
CANDIDATE_TABLE_COLUMN_EMAIL_VALIDATED = "EmailValidated"
CANDIDATE_TABLE_COLUMN_MOBILE_VALIDATED = "MobileValidated"
CANDIDATE_TABLE_COLUMN_INDEXED = "Indexed"
CANDIDATE_TABLE_COLUMN_DISABLED = "Disabled"
CANDIDATE_TABLE_COLUMN_PROFILE_STATUS = "ProfileStatus"

CANDIDATE_TABLE_COLUMN_RESUME_TYPE = "ResumeType"
CANDIDATE_TABLE_COLUMN_RESUME_ID = "ResumeId"
CANDIDATE_TABLE_COLUMN_RESUME_NAME = "ResumeName"
CANDIDATE_TABLE_COLUMN_RESUME_PATH = "ResumePath"
CANDIDATE_TABLE_COLUMN_RESUME_PARSED = "ResumeParsed"
CANDIDATE_TABLE_COLUMN_RESUME_URL = "Url"

CANDIDATE_SKILL_TABLE_NAME = "candidateskill"
CANDIDATE_SKILL_TABLE_CANDIDATE_ID = "CandidateId"
CANDIDATE_SKILL_TABLE_SKILL_ID = "SkillId"
CANDIDATE_SKILL_TABLE_EXPERIENCE = "Experience"
CANDIDATE_SKILL_TABLE_ID = "Id"
CANDIDATE_SKILL_TABLE_LAST_UPDATED_TIME = "LastUpdatedTime"
CANDIDATE_SKILL_TABLE_RATING = "Rating"

CANDIDATE_JOB_SKILL_TABLE_NAME = "candidatejobskill"
CANDIDATE_JOB_SKILL_TABLE_ID = "Id"
CANDIDATE_JOB_SKILL_TABLE_CANDIDATE_ID = "CandidateId"
CANDIDATE_JOB_SKILL_TABLE_JOB_ID = "JobId"
CANDIDATE_JOB_SKILL_TABLE_LAST_UPDATED_TIME = "LastUpdatedTime"
CANDIDATE_JOB_SKILL_TABLE_CREATED_TIME = "CreatedTime"
CANDIDATE_JOB_SKILL_TABLE_CANDIDATE_INTERESTED = "CandidateIntrested"
CANDIDATE_JOB_SKILL_TABLE_EMPLOYER_INTERESTED = "EmployerIntrested"
CANDIDATE_JOB_SKILL_TABLE_TOTAL_SCORE = "TotalScore"
CANDIDATE_JOB_SKILL_TABLE_STATE = "State"

JOB_TABLE_NAME = "jobs"
JOB_TABLE_ID = "Id"
JOB_TABLE_INDEX = "Indexed"
JOB_TABLE_LAST_UPDATED_TIME = "LastUpdateTime"


########################################################################################################################
class CsvFileValidation:
    def __init__(self, path):
        self.path = path
        self.rows = []
        self.fields = []

        self.csv_file_reading()

    # reading csv file
    def csv_file_reading(self):
        with open(self.path, 'r') as csvfile:
            # creating a csv reader object
            csvreader = csv.reader(csvfile)

            # extracting field names through first row
            self.fields = next(csvreader)

            # extracting each data row one by one
            for row in csvreader:
                self.rows.append(row)

            # get total number of rows
            logging.info("Total no. of rows: %d" % csvreader.line_num)

    def csv_required_details(self):
        candidate_details = []
        candidate_skill_id = []
        resume_list = []
        for row in self.rows:
            # parsing each column of a row
            count = 0
            candidate_detail_list = []
            skill_id_list = []
            resume_link_list = []
            if len(row) == 0:
                continue
            for col in row:
                if self.fields[count] == FIELD_NAMES[0]:  # First Name
                    try:
                        if len(col) > 0 or col is not None:
                            candidate_detail_list.append(col)
                        else:
                            candidate_detail_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading first name {e}")
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[1]:  # Last Name
                    try:
                        if len(col) > 0 or col is not None:
                            candidate_detail_list.append(col)
                        else:
                            candidate_detail_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading last name {e}")
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[2]:  # Mob No
                    try:
                        if len(col) > 0 or col is not None:
                            mob_no = col
                            mob_no = "+91." + str(mob_no)
                            candidate_detail_list.append(mob_no)
                            skill_id_list.append(mob_no)
                            resume_link_list.append(mob_no)
                        else:
                            candidate_detail_list.append('0')
                            skill_id_list.append('0')
                            resume_link_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Mobile Number {e}")
                        candidate_detail_list.append('0')
                        skill_id_list.append('0')
                        resume_link_list.append('0')
                if self.fields[count] == FIELD_NAMES[3]:  # Email Id
                    try:
                        if len(col) > 0 or col is not None:
                            candidate_detail_list.append(col)
                        else:
                            candidate_detail_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Email Id {e}")
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[4]:  # Created time
                    try:
                        if len(col) > 0 or col is not None:
                            created_time = col
                            try:
                                created_time = created_time.replace("  ", " ")
                            except Exception as e:
                                pass
                            try:
                                created_time = datetime.strptime(created_time, '%d-%b-%Y %H:%M:%S')
                            except Exception as e:
                                logging.exception(f"{e}")
                                try:
                                    created_time = datetime.strptime(created_time, '%d-%m-%Y %H:%M:%S')
                                except Exception as e:
                                    logging.exception(f"{e}")
                                    try:
                                        created_time = datetime.strptime(created_time, '%d-%m-%Y %H:%M')
                                    except Exception as e:
                                        logging.exception(f"{e}")
                                        try:
                                            created_time = datetime.strptime(created_time, '%d-%b-%Y %H:%M')
                                        except Exception as e:
                                            logging.exception(f"{e}")
                            candidate_detail_list.append(created_time)
                            print(created_time)
                            print(type(created_time))
                        else:
                            candidate_detail_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Created Date {e}")
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[5]:  # Test Score
                    try:
                        if len(col) > 0 or col is not None:
                            skill_id_list.append(float(col))
                        else:
                            skill_id_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Test Score {e}")
                        skill_id_list.append('0')
                if self.fields[count] == FIELD_NAMES[7]:  # Notice Period
                    try:
                        if len(col) > 0 or col is not None:
                            candidate_detail_list.append(float(col))
                        else:
                            candidate_detail_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Notice Period {e}")
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[8]:  # Current CTC
                    try:
                        if len(col) > 0 or col is not None:
                            candidate_detail_list.append(float(col))
                        else:
                            candidate_detail_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Current CTC {e}")
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[9]:  # Expected CTC
                    try:
                        if len(col) > 0 or col is not None:
                            candidate_detail_list.append(float(col))
                        else:
                            candidate_detail_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Expected CTC {e}")
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[11]:  # Skill1 ID
                    try:
                        if len(col) > 0 or col is not None:
                            skill_id_list.append(int(col))
                        else:
                            skill_id_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Skill1 ID {e}")
                        skill_id_list.append('0')
                if self.fields[count] == FIELD_NAMES[12]:  # Skill1 Years
                    try:
                        if len(col) > 0 or col is not None:
                            skill_id_list.append(float(col))
                        else:
                            skill_id_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Skill1 Years {e}")
                        skill_id_list.append('0')
                if self.fields[count] == FIELD_NAMES[14]:  # Skill2 ID
                    try:
                        if len(col) > 0 or col is not None:
                            skill_id_list.append(int(col))
                        else:
                            skill_id_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Skill2 ID {e}")
                        skill_id_list.append('0')
                if self.fields[count] == FIELD_NAMES[15]:  # Skill2 Years
                    try:
                        if len(col) > 0 or col is not None:
                            skill_id_list.append(float(col))
                        else:
                            skill_id_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Skill2 Years {e}")
                        skill_id_list.append('0')
                if self.fields[count] == FIELD_NAMES[17]:  # Skill3 ID
                    try:
                        if len(col) > 0 or col is not None:
                            skill_id_list.append(int(col))
                        else:
                            skill_id_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Skill3 ID {e}")
                        skill_id_list.append('0')
                if self.fields[count] == FIELD_NAMES[18]:  # Skill3 Years
                    try:
                        if len(col) > 0 or col is not None:
                            skill_id_list.append(float(col))
                        else:
                            skill_id_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Skill3 Years {e}")
                        skill_id_list.append('0')
                if self.fields[count] == FIELD_NAMES[6]:  # Resume Link
                    try:
                        if len(col) > 0 or col is not None:
                            resume_link_list.append(col)
                        else:
                            resume_link_list.append('0')
                    except Exception as e:
                        logging.exception(f"Error reading Resume Link {e}")
                        resume_link_list.append('0')
                if count < len(self.fields) - 1:
                    count += 1

            candidate_details.append(candidate_detail_list)
            candidate_skill_id.append(skill_id_list)
            resume_list.append(resume_link_list)

        return candidate_details, candidate_skill_id, resume_list

########################################################################################################################


########################################################################################################################
class ExcelFileValidation:
    def __init__(self, path):
        self.path = path
        self.sheet_obj = None
        self.row = 0
        self.column = 0
        self.excel_success_row_list = []
        self.excel_success_row_list.append(FIELD_NAMES)
        self.excel_error_row_list = []
        self.excel_error_row_list.append(FIELD_NAMES)

        self.connection = create_server_connection()

        self.excel_file_reading()

    def excel_file_reading(self):
        # To open the workbook
        # workbook object is created
        wb_obj = openpyxl.load_workbook(self.path)

        # Get workbook active sheet object
        # from the active attribute
        self.sheet_obj = wb_obj.active

        # Getting the value of maximum rows
        # and column
        self.row = self.sheet_obj.max_row
        self.column = self.sheet_obj.max_column

    def excel_required_details(self):
        candidate_details = []
        candidate_skill_id = []
        resume_list = []
        for j in range(3, self.row + 1):
            candidate_detail_list = []
            skill_id_list = []
            resume_link_list = []
            for i in range(1, self.column + 1):
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[0]:
                    try:
                        name = self.sheet_obj.cell(row=j, column=i).value
                        if name is None or type(name) != str:
                            candidate_detail_list.append('0')
                        else:
                            candidate_detail_list.append(name)
                    except Exception as e:
                        logging.exception(f"Error reading first name {e}")
                        candidate_detail_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[1]:
                    try:
                        last_name = self.sheet_obj.cell(row=j, column=i).value
                        if last_name is None or type(last_name) != str:
                            candidate_detail_list.append('0')
                        else:
                            candidate_detail_list.append(last_name)
                    except Exception as e:
                        logging.exception(f"Error reading last name {e}")
                        candidate_detail_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[2]:
                    try:
                        mob_no = self.sheet_obj.cell(row=j, column=i).value
                        if mob_no is None or type(mob_no) != int or len(str(mob_no)) < 10:
                            candidate_detail_list.append('0')
                            skill_id_list.append('0')
                            resume_link_list.append('0')
                        else:
                            mob_no = "+91." + str(mob_no)
                            candidate_detail_list.append(mob_no)
                            skill_id_list.append(mob_no)
                            resume_link_list.append(mob_no)
                    except Exception as e:
                        logging.exception(f"Error reading Mobile Number {e}")
                        candidate_detail_list.append('0')
                        skill_id_list.append('0')
                        resume_link_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[3]:
                    try:
                        email_id = self.sheet_obj.cell(row=j, column=i).value
                        if email_id is None:
                            candidate_detail_list.append('0')
                        else:
                            candidate_detail_list.append(email_id)
                    except Exception as e:
                        logging.exception(f"Error reading Email Id {e}")
                        candidate_detail_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[4]:
                    try:
                        created_time = self.sheet_obj.cell(row=j, column=i).value
                        if created_time is None:
                            candidate_detail_list.append('0')
                        else:
                            logging.warning(f"created_time = {created_time} \n created_time Type = {type(created_time)}")
                            created_time = created_time.strptime(str(created_time), '%Y-%m-%d %H:%M:%S.%f')
                            candidate_detail_list.append(created_time)
                    except Exception as e:
                        logging.exception(f"Error reading Created Date {e}")
                        candidate_detail_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[5]:
                    try:
                        test_score = self.sheet_obj.cell(row=j, column=i).value
                        if test_score is None or (type(test_score) != float and type(test_score) != int):
                            skill_id_list.append('0')
                        else:
                            skill_id_list.append(test_score)
                    except Exception as e:
                        logging.exception(f"Error reading Test Score {e}")
                        skill_id_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[7]:
                    try:
                        notice_period = self.sheet_obj.cell(row=j, column=i).value
                        if notice_period is None or (type(notice_period) != float and type(notice_period) != int):
                            candidate_detail_list.append('0')
                        else:
                            candidate_detail_list.append(notice_period)
                    except Exception as e:
                        logging.exception(f"Error reading Notice Period {e}")
                        candidate_detail_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[8]:
                    try:
                        current_ctc = self.sheet_obj.cell(row=j, column=i).value
                        if current_ctc is None or (type(current_ctc) != float and type(current_ctc) != int):
                            candidate_detail_list.append('0')
                        else:
                            candidate_detail_list.append(current_ctc)
                    except Exception as e:
                        logging.exception(f"Error reading Current CTC {e}")
                        candidate_detail_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[9]:
                    try:
                        expected_ctc = self.sheet_obj.cell(row=j, column=i).value
                        if expected_ctc is None or (type(expected_ctc) != float and type(expected_ctc) != int):
                            candidate_detail_list.append('0')
                        else:
                            candidate_detail_list.append(expected_ctc)
                    except Exception as e:
                        logging.exception(f"Error reading Expected CTC {e}")
                        candidate_detail_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[11]:
                    try:
                        skill1_id = self.sheet_obj.cell(row=j, column=i).value
                        if skill1_id is None or  type(skill1_id) != int:
                            skill_id_list.append('0')
                        else:
                            skill_id_list.append(skill1_id)
                    except Exception as e:
                        logging.exception(f"Error reading Skill1 Id {e}")
                        skill_id_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[12]:
                    try:
                        skill1_year = self.sheet_obj.cell(row=j, column=i).value
                        if skill1_year is None or (type(skill1_year) != float and type(skill1_year) != int):
                            skill_id_list.append('0')
                        else:
                            skill_id_list.append(skill1_year)
                    except Exception as e:
                        logging.exception(f"Error reading Skill1 Years {e}")
                        skill_id_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[14]:
                    try:
                        skill2_id = self.sheet_obj.cell(row=j, column=i).value
                        if skill2_id is None or type(skill2_id) != int:
                            skill_id_list.append('0')
                        else:
                            skill_id_list.append(skill2_id)
                    except Exception as e:
                        logging.exception(f"Error reading Skill2 Id {e}")
                        skill_id_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[15]:
                    try:
                        skill2_year = self.sheet_obj.cell(row=j, column=i).value
                        if skill2_year is None or (type(skill2_year) != float and type(skill2_year) != int):
                            skill_id_list.append('0')
                        else:
                            skill_id_list.append(skill2_year)
                    except Exception as e:
                        logging.exception(f"Error reading Skill2 Years {e}")
                        skill_id_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[17]:
                    try:
                        skill3_id = self.sheet_obj.cell(row=j, column=i).value
                        if skill3_id is None or type(skill3_id) != int:
                            skill_id_list.append('0')
                        else:
                            skill_id_list.append(skill3_id)
                    except Exception as e:
                        logging.exception(f"Error reading Skill3 Id {e}")
                        skill_id_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[18]:
                    try:
                        skill3_year = self.sheet_obj.cell(row=j, column=i).value
                        if skill3_year is None or (type(skill3_year) != float and type(skill3_year) != int) == 0:
                            skill_id_list.append('0')
                        else:
                            skill_id_list.append(skill3_year)
                    except Exception as e:
                        logging.exception(f"Error reading Skill3 Years {e}")
                        skill_id_list.append('0')
                if self.sheet_obj.cell(row=2, column=i).value == FIELD_NAMES[6]:
                    try:
                        url = self.sheet_obj.cell(row=j, column=i).value
                        if url is None:
                            logging.warning("URL not given")
                            resume_link_list.append('0')
                        else:
                            resume_link_list.append(url)
                    except Exception as e:
                        logging.exception(f"Error reading Resume Link {e}")
                        resume_link_list.append('0')

            candidate_details.append(candidate_detail_list)
            candidate_skill_id.append(skill_id_list)
            resume_list.append(resume_link_list)

        return candidate_details, candidate_skill_id, resume_list

########################################################################################################################


def reading_config_file():
    f = open(CONFIG_FILE_PATH, "r")
    data = json.load(f)
    global HOST_NAME
    HOST_NAME = data["host_name"]
    global USER_NAME
    USER_NAME = data["user_name"]
    global PASSWORD
    PASSWORD = data["password"]
    global DB_NAME
    DB_NAME = data["db_name"]
    global RESUME_DIRECTORY_PATH
    RESUME_DIRECTORY_PATH = data["resume_directory_path"]


def create_server_connection():
    connect = None
    try:
        connect = mysql.connector.connect(
            host=HOST_NAME,
            user=USER_NAME,
            passwd=PASSWORD,
            database=DB_NAME
        )
        logging.info("Connection Successful")
    except Error as err:
        logging.error(f"Error: {err}")
    return connect


def execute_query(connect, query):
    cursor = connect.cursor()
    try:
        cursor.execute(query)
        connect.commit()
        logging.info("Query Execute Successfully")
        return True
    except Error as err:
        logging.error(f"Error : {err}")
        return False


def read_query(connect, query):
    cursor = connect.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()

    except Error as err:
        logging.error(f"Error: {err}")
    return result


def update_candidate_table(candidate_details):
    execution_flag = False
    connection = create_server_connection()
    logging.warning(candidate_details)
    try:
        for detail in candidate_details:
            logging.warning(detail)
            query = f'''select {CANDIDATE_TABLE_COLUMN_ID}, {CANDIDATE_TABLE_COLUMN_EMAIL_ID}, 
            {CANDIDATE_TABLE_DISPLAY_NAME} 
            From {CANDIDATE_TABLE_NAME} where {CANDIDATE_TABLE_COLUMN_MOB_NO} = {detail[2]}'''

            detail_result = read_query(connection, query)

            time_now = datetime.now()
            time_now = time_now.strftime('%Y-%m-%d %H:%M:%S')

            milliseconds = int(time.time() * 1000)
            display_name = detail[0] + detail[1] + str(milliseconds)

            if detail_result is not None and len(detail_result) > 0:
                update_query = f'''UPDATE {CANDIDATE_TABLE_NAME} SET `{CANDIDATE_TABLE_COLUMN_CCTC}` = {detail[6]}, 
                `{CANDIDATE_TABLE_COLUMN_ECTC}` = {detail[7]}, `{CANDIDATE_TABLE_FIRST_NAME}` = '{detail[0]}', 
                `{CANDIDATE_TABLE_LAST_NAME}` = '{detail[1]}', `{CANDIDATE_TABLE_COLUMN_NOTICE_PERIOD}` = {detail[5]},
                `{CANDIDATE_TABLE_COLUMN_LAST_UPDATED_TIME}` = '{time_now}',
                `{CANDIDATE_TABLE_COLUMN_EMAIL_VALIDATED}` = {1}, `{CANDIDATE_TABLE_COLUMN_MOBILE_VALIDATED}` = {1},
                `{CANDIDATE_TABLE_COLUMN_DISABLED}` = {0}, `{CANDIDATE_TABLE_COLUMN_INDEXED}` = {0},
                `{CANDIDATE_TABLE_COLUMN_PROFILE_STATUS}` = {20}
                WHERE ({CANDIDATE_TABLE_COLUMN_ID} = {detail_result[0][0]});'''

                execution_flag = execute_query(connection, update_query)

                email_update_flag = False
                if detail_result[0][1] != detail[3]:
                    update_query = f'''UPDATE {CANDIDATE_TABLE_NAME} SET
                    `{CANDIDATE_TABLE_COLUMN_EMAIL_ID}` = '{detail[3]}'
                    WHERE ({CANDIDATE_TABLE_COLUMN_ID} = {detail_result[0][0]});'''

                    email_update_flag = execute_query(connection, update_query)

                print(f"Display Name : {detail_result[0][2]}")
                display_name_flag = False
                if detail_result[0][2] is None:
                    update_query = f'''UPDATE {CANDIDATE_TABLE_NAME} SET
                    `{CANDIDATE_TABLE_DISPLAY_NAME}` = '{display_name}'
                    WHERE ({CANDIDATE_TABLE_COLUMN_ID} = {detail_result[0][0]});'''

                    display_name_flag = execute_query(connection, update_query)

            else:
                detail[4] = detail[4].strftime('%Y-%m-%d %H:%M:%S')
                insert_query = f'''INSERT INTO {CANDIDATE_TABLE_NAME} (`{CANDIDATE_TABLE_COLUMN_CCTC}`, 
                `{CANDIDATE_TABLE_COLUMN_ECTC}`, `{CANDIDATE_TABLE_COLUMN_EMAIL_ID}`, `{CANDIDATE_TABLE_FIRST_NAME}`, 
                `{CANDIDATE_TABLE_LAST_NAME}`, `{CANDIDATE_TABLE_COLUMN_MOB_NO}`, `{CANDIDATE_TABLE_DISPLAY_NAME}`,
                `{CANDIDATE_TABLE_COLUMN_NOTICE_PERIOD}`, `{CANDIDATE_TABLE_COLUMN_CREATED_TIME}`,
                `{CANDIDATE_TABLE_COLUMN_LAST_UPDATED_TIME}`, `{CANDIDATE_TABLE_COLUMN_EMAIL_VALIDATED}`,
                `{CANDIDATE_TABLE_COLUMN_MOBILE_VALIDATED}`, `{CANDIDATE_TABLE_COLUMN_DISABLED}`, 
                `{CANDIDATE_TABLE_COLUMN_INDEXED}`, `{CANDIDATE_TABLE_COLUMN_PROFILE_STATUS}`) 
                VALUES ({detail[6]}, {detail[7]}, '{detail[3]}', '{detail[0]}', '{detail[1]}', {detail[2]}, 
                '{display_name}', {detail[5]}, '{detail[4]}', '{time_now}', {1}, {1}, {0}, {0}, {20});'''

                execution_flag = execute_query(connection, insert_query)

        connection.close()
        if execution_flag:
            return True
        else:
            return False
    except Exception as e:
        logging.exception(f"Error updating candidate table {e}")
        return False


def candidate_skill_table_update(candidate_skills, job_id):
    print(candidate_skills)
    # candidate_skills = [0:mob no, 1:test_score, 2:skill1_id, 3:skill1_year,
    # 4:skill2_id, 5:skill2_year, 6:skill3_id, 7:skill3_year,]
    execution_flag = False
    connection = create_server_connection()
    for candidate_skill in candidate_skills:
        id_query = f'''select {CANDIDATE_TABLE_COLUMN_ID} From {CANDIDATE_TABLE_NAME} 
        where {CANDIDATE_TABLE_COLUMN_MOB_NO} = {candidate_skill[0]}'''
        id_result = read_query(connection, id_query)

        if id_result is not None and len(id_result) > 0:
            time_now = datetime.now()
            time_now = time_now.strftime('%Y-%m-%d %H:%M:%S')

            for i in range(0, 5, 2):
                if candidate_skill[i+2] == '0':
                    continue
                skill_check_query = f'''select {CANDIDATE_SKILL_TABLE_ID} from {CANDIDATE_SKILL_TABLE_NAME} 
                where {CANDIDATE_SKILL_TABLE_CANDIDATE_ID} = {int(id_result[0][0])} and 
                {CANDIDATE_SKILL_TABLE_SKILL_ID} = {candidate_skill[i+2]}'''

                skill_check_result = read_query(connection, skill_check_query)

                if skill_check_result is not None and len(skill_check_result) > 0:
                    logging.info(f"Skill id {candidate_skill[i+2]} \
                    already present for candidate id {int(id_result[0][0])}")

                    skill_update_query = f'''UPDATE `{CANDIDATE_SKILL_TABLE_NAME}` 
                    SET `{CANDIDATE_SKILL_TABLE_EXPERIENCE}` = {candidate_skill[i+3]},
                    `{CANDIDATE_SKILL_TABLE_LAST_UPDATED_TIME}` = '{time_now}',
                    `{CANDIDATE_SKILL_TABLE_RATING}` = {0}
                    WHERE ({CANDIDATE_SKILL_TABLE_ID} = {int(skill_check_result[0][0])});'''

                    skill_update_flag = execute_query(connection, skill_update_query)

                else:
                    skill_insert_query = f'''INSERT INTO {CANDIDATE_SKILL_TABLE_NAME} 
                    (`{CANDIDATE_SKILL_TABLE_CANDIDATE_ID}`, `{CANDIDATE_SKILL_TABLE_SKILL_ID}`, 
                    `{CANDIDATE_SKILL_TABLE_EXPERIENCE}`, `{CANDIDATE_SKILL_TABLE_LAST_UPDATED_TIME}`,
                    `{CANDIDATE_SKILL_TABLE_RATING}`) 
                    VALUES ({int(id_result[0][0])}, {candidate_skill[i+2]}, {candidate_skill[i+3]}, '{time_now}'
                    , {0});'''

                    execution_flag = execute_query(connection, skill_insert_query)

            cjs_check_query = f'''SELECT `{CANDIDATE_JOB_SKILL_TABLE_ID}` FROM {CANDIDATE_JOB_SKILL_TABLE_NAME} 
                            WHERE (`{CANDIDATE_JOB_SKILL_TABLE_CANDIDATE_ID}` = {int(id_result[0][0])} and 
                            `{CANDIDATE_JOB_SKILL_TABLE_JOB_ID}` = {job_id});'''

            cjs_check_result = read_query(connection, cjs_check_query)

            if cjs_check_result is not None and len(cjs_check_result) > 0:
                logging.info(f"job id {job_id} already present for candidate id {int(id_result[0][0])}")

                update_cjs_query = f'''UPDATE `{CANDIDATE_JOB_SKILL_TABLE_NAME}` 
                                SET `{CANDIDATE_JOB_SKILL_TABLE_LAST_UPDATED_TIME}` = '{time_now}',
                                `{CANDIDATE_JOB_SKILL_TABLE_CANDIDATE_INTERESTED}` = {1},
                                `{CANDIDATE_JOB_SKILL_TABLE_EMPLOYER_INTERESTED}` = {1},
                                `{CANDIDATE_JOB_SKILL_TABLE_TOTAL_SCORE}` = {candidate_skill[1]},
                                `{CANDIDATE_JOB_SKILL_TABLE_STATE}` = '{"Apply"}'
                                WHERE (`{CANDIDATE_JOB_SKILL_TABLE_ID}` = {int(cjs_check_result[0][0])});'''

                update_cjs_flag = execute_query(connection, update_cjs_query)

            else:
                insert_cjs_query = f'''INSERT INTO {CANDIDATE_JOB_SKILL_TABLE_NAME} 
                                (`{CANDIDATE_JOB_SKILL_TABLE_CANDIDATE_ID}`, `{CANDIDATE_JOB_SKILL_TABLE_JOB_ID}`, 
                                `{CANDIDATE_JOB_SKILL_TABLE_CREATED_TIME}`, 
                                `{CANDIDATE_JOB_SKILL_TABLE_LAST_UPDATED_TIME}`
                                `{CANDIDATE_JOB_SKILL_TABLE_CANDIDATE_INTERESTED}`,
                                `{CANDIDATE_JOB_SKILL_TABLE_EMPLOYER_INTERESTED}`,
                                `{CANDIDATE_JOB_SKILL_TABLE_TOTAL_SCORE}`, `{CANDIDATE_JOB_SKILL_TABLE_STATE}`) 
                                VALUES ({int(id_result[0][0])}, {job_id}, '{time_now}', '{time_now}', 
                                {1}, {1}, {candidate_skill[1]}, '{"Apply"}');'''

                insert_cjs_flag = execute_query(connection, insert_cjs_query)

        else:
            logging.warning("Invalid candidate Id")
    connection.close()


def resume_download(resume_link_list):
    connection = create_server_connection()
    try:
        if os.path.isdir(RESUME_DIRECTORY_PATH):
            pass
        else:
            try:
                os.mkdir(RESUME_DIRECTORY_PATH)
            except Exception as e:
                logging.exception(f"Error creating parent directory {RESUME_DIRECTORY_PATH}: {e}")
                connection.close()
                return

        for link_detail in resume_link_list:
            print(link_detail)
            query = f'''select {CANDIDATE_TABLE_COLUMN_ID} From {CANDIDATE_TABLE_NAME} 
            where {CANDIDATE_TABLE_COLUMN_MOB_NO} = {link_detail[0]}'''

            detail_result = read_query(connection, query)

            if detail_result is not None and len(detail_result) > 0:
                candidate_id = detail_result[0][0]
                if link_detail[1] == '0':
                    logging.warning(f"resume link not available for candidate id: {candidate_id}")
                    continue

                resume_url = link_detail[1]
                update_query = f'''UPDATE {CANDIDATE_TABLE_NAME} SET 
                                `{CANDIDATE_TABLE_COLUMN_RESUME_URL}` = '{resume_url}'
                                WHERE ({CANDIDATE_TABLE_COLUMN_ID} = {candidate_id});'''

                update_query_flag = execute_query(connection, update_query)

                print(f"update flag: {update_query_flag}")

                if not update_query_flag:
                    logging.warning(f"Error Updating URL column of candidate table")

                resume_directory_path = RESUME_DIRECTORY_PATH + str(candidate_id) + "/"
                if os.path.isdir(resume_directory_path):
                    pass
                else:
                    try:
                        os.mkdir(resume_directory_path)
                    except Exception as e:
                        logging.exception(f"Error creating parent directory {resume_directory_path}: {e}")
                        continue

                file_id = resume_url.split("/")[5]

                URL = "https://docs.google.com/uc?export=download"

                session = requests.Session()

                response = session.get(URL, params={'id': file_id, 'confirm': 1}, stream=True)
                token = get_confirm_token(response)

                if token:
                    params = {'id': file_id, 'confirm': token}
                    response = session.get(URL, params=params, stream=True)
                filename = response.headers['Content-Disposition'].split(';')[1]
                filename = filename.split('\"')[1]

                destination = resume_directory_path + filename

                file_save_flag = save_response_content(response, destination)

                if file_save_flag:
                    update_query = f'''UPDATE {CANDIDATE_TABLE_NAME} SET 
                    `{CANDIDATE_TABLE_COLUMN_RESUME_PATH}` = '{destination}',
                    `{CANDIDATE_TABLE_COLUMN_RESUME_NAME}` = '{filename}',
                    `{CANDIDATE_TABLE_COLUMN_RESUME_TYPE}` = '{filename.split('.')[-1]}'
                    WHERE ({CANDIDATE_TABLE_COLUMN_ID} = {candidate_id});'''

                    update_query_flag = execute_query(connection, update_query)
                else:
                    logging.warning(f"Resume file does not saved for candidate id {candidate_id}")

            else:
                logging.warning(f"Candidate id not found against mob no. {link_detail[0]}")
        connection.close()
    except Exception as e:
        logging.exception(f"Error downloading resume {e}")
        connection.close()


def get_confirm_token(response):
    for key, value in response.cookies.items():
        if key.startswith('download_warning'):
            return value

    return None


def save_response_content(response, destination):
    try:
        chunk_size = 32768

        with open(destination, "wb") as f:
            for chunk in response.iter_content(chunk_size):
                if chunk:  # filter out keep-alive new chunks
                    f.write(chunk)
        return True
    except Exception as e:
        logging.exception(f"Error saving resume file {e}")
        return False


def main():
    reading_config_file()
    while True:
        try:
            connection = create_server_connection()

            query = f'''
            select {TABLE_COLUMN_ID}, {TABLE_COLUMN_PATH}, {TABLE_COLUMN_JOB_ID} From {TABLE_NAME} 
            where {TABLE_COLUMN_PROCESS_STATE} = '1';
            '''

            query_results = read_query(connection, query)

            if query_results is not None and len(query_results) > 0:
                for result in query_results:
                    row_id = result[0]
                    path = str(result[1])
                    print(f"Working on {path}")
                    job_id = int(result[2])

                    candidate_details = []
                    candidate_skill_id = []
                    resume_link_list = []

                    if ".csv" in path:
                        path = path.replace('.csv', '-Success.csv')
                        csv_obj = CsvFileValidation(path)
                        candidate_details, candidate_skill_id, resume_link_list = csv_obj.csv_required_details()
                    elif ".xlsx" in path:
                        path = path.replace('.xlsx', '-Success.xlsx')
                        excel_obj = ExcelFileValidation(path)
                        candidate_details, candidate_skill_id, resume_link_list = excel_obj.excel_required_details()
                    else:
                        logging.warning(f"Given File Path for row-id {row_id} is invalid")
                    if len(candidate_details) > 0 and len(candidate_skill_id) > 0:
                        update_table_flag = update_candidate_table(candidate_details)
                        if update_table_flag:
                            candidate_skill_table_update(candidate_skill_id, job_id)
                            resume_download(resume_link_list)

                            time_now = datetime.now()
                            time_now = time_now.strftime('%Y-%m-%d %H:%M:%S')

                            query = f'''UPDATE {JOB_TABLE_NAME}
                                        SET {JOB_TABLE_INDEX} = {0},
                                        {JOB_TABLE_LAST_UPDATED_TIME} = '{time_now}'
                                        WHERE {JOB_TABLE_ID} = '{job_id}';'''
                            execute_query(connection, query)

                            query = f'''UPDATE {TABLE_NAME}
                                        SET {TABLE_COLUMN_PROCESS_STATE} = '2'
                                        WHERE {TABLE_COLUMN_ID} = '{row_id}';'''
                            execute_query(connection, query)
                        else:
                            logging.warning(f"Entries from {path} are not updated in the tables")
                            print(f"Entries from {path} are not updated in the tables")

                    print("Sleeping for 10 sec")
                    logging.info("Sleeping for 10 sec")
                    time.sleep(10)
            else:
                logging.info("No new file to update")
                print("No new file to update")
            connection.close()
        except Exception as e:
            logging.exception(f"error in main method {e}")
        print("Sleeping for 15 sec")
        logging.info("Sleeping for 15 sec")
        time.sleep(15)


if __name__ == "__main__":
    main()
