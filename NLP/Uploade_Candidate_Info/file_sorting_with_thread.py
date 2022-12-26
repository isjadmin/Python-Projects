import openpyxl
from openpyxl import Workbook
import re
from datetime import datetime
import mysql.connector
from mysql.connector import Error
import csv
import logging
import threading
import json
import time


FIELD_NAMES = ['first name', 'last name', 'mob no', 'email', 'created time', 'test score', 'linked resume',
               'notice period', 'current ctc', 'expected ctc', 'skill1', 'skill1 ID', 'skill1 years', 'skill2',
               'skill2 ID', 'skill2 years', 'skill3', 'skill3 ID', 'skill3 years', 'remark']

CONFIG_FILE_PATH = "config.json"

HOST_NAME = ""
USER_NAME = ""
PASSWORD = ""
DB_NAME = ""

TABLE_NAME = "upload"
TABLE_COLUMN_ID = "Id"
TABLE_COLUMN_PATH = "Path"
TABLE_COLUMN_PROCESS_STATE = "ProcessedState"
TABLE_COLUMN_FILE_UPDATE_TIME = "FileUpdateTime"
TABLE_COLUMN_JOB_ID = "JobId"
TABLE_COLUMN_SUCCESS_FILE_PATH = "SuccessPath"
TABLE_COLUMN_ERROR_FILE_PATH = "FailurePath"
TABLE_COLUMN_SUCCESS_FILE_ROWS = "SuccessCount"
TABLE_COLUMN_ERROR_FILE_ROWS = "FailureCount"

TABLE_JOB_SKILL_NAME = "jobskill"
TABLE_JOB_SKILL_JOB_ID = "JobId"
TABLE_JOB_SKILL_SKILL_NAME = "SkillName"
TABLE_JOB_SKILL_SKILL_ID = "SkillId"

TABLE_SKILL_NAME = "skills"
TABLE_SKILL_TITLE = "Title"
TABLE_SKILL_ID = "Id"


########################################################################################################################
class CsvFileValidation:
    def __init__(self, path, job_skills):
        self.path = path
        self.rows = []
        self.fields = []
        self.csv_success_row_list = []
        self.csv_error_row_list = []
        self.csv_success_file_name = self.path.replace('.csv', '-Success.csv')
        self.csv_error_file_name = self.path.replace('.csv', '-Error.csv')
        self.job_skills = job_skills

        self.connection = create_server_connection()

        self.csv_file_reading()
        self.csv_file_validation()

    # reading csv file
    def csv_file_reading(self):
        try:
            with open(self.path, 'r') as csvfile:
                # creating a csv reader object
                csvreader = csv.reader(csvfile)

                # extracting field names through first row
                self.fields = next(csvreader)

                # extracting each data row one by one
                for row in csvreader:
                    self.rows.append(row)

                # get total number of rows
                logging.info("Total no. of rows: %d" % csvreader.line_num)
                print("Total no. of rows: %d" % csvreader.line_num)
        except Exception as e:
            logging.exception(f"Error Reading {self.path} file {e}")

    def csv_file_validation(self):
        for row in self.rows:
            # parsing each column of a row
            count = 0
            csv_success_flag = True
            remark = ""
            skill_list = []
            skill1_id = None
            skill2_id = None
            skill3_id = None
            if len(row) == 0:
                continue
            for col in row:
                if self.fields[count] == FIELD_NAMES[0]:
                    try:
                        if len(col) > 0:
                            pass
                        else:
                            csv_success_flag = False
                            remark += "Invalid First Name, "
                    except Exception as e:
                        logging.exception(f"Error getting first name {e}")
                        csv_success_flag = False
                        remark += "Invalid First Name, "
                if self.fields[count] == FIELD_NAMES[1]:
                    try:
                        if len(col) > 0:
                            pass
                        else:
                            csv_success_flag = False
                            remark += "Invalid Last Name, "
                    except Exception as e:
                        logging.exception(f"Error getting last name {e}")
                        csv_success_flag = False
                        remark += "Invalid Last Name, "
                if self.fields[count] == FIELD_NAMES[2]:
                    try:
                        if len(col) < 10:
                            csv_success_flag = False
                            remark += "Invalid Mobile No., "
                        else:
                            pass
                    except Exception as e:
                        logging.exception(f"{e}")
                        csv_success_flag = False
                        remark += "Invalid Mobile No., "
                if self.fields[count] == FIELD_NAMES[3]:
                    try:
                        if email_validation(col):
                            pass
                        else:
                            csv_success_flag = False
                            remark += "Invalid Email-ID, "
                    except Exception as e:
                        logging.exception(f"Error getting Email Id {e}")
                        csv_success_flag = False
                        remark += "Invalid Email-ID, "
                if self.fields[count] == FIELD_NAMES[4]:
                    try:
                        date_time = datetime.strptime(col, '%d-%b-%Y %H:%M:%S')
                        if str(type(date_time)) == "<class 'datetime.datetime'>":
                            pass
                        else:
                            print("in else")
                            csv_success_flag = False
                            remark += "Invalid Created Date, "
                    except Exception as e:
                        logging.exception(f"{e}")
                        try:
                            date_time = datetime.strptime(col, '%d-%m-%Y %H:%M:%S')
                            if str(type(date_time)) == "<class 'datetime.datetime'>":
                                pass
                            else:
                                print("in else")
                                csv_success_flag = False
                                remark += "Invalid Created Date, "
                        except Exception as e:
                            logging.exception(f"{e}")
                            try:
                                date_time = datetime.strptime(col, '%d-%b-%Y %H:%M')
                                if str(type(date_time)) == "<class 'datetime.datetime'>":
                                    pass
                                else:
                                    print("in else")
                                    csv_success_flag = False
                                    remark += "Invalid Created Date, "
                            except Exception as e:
                                logging.exception(f"{e}")
                                try:
                                    date_time = datetime.strptime(col, '%d-%m-%Y %H:%M')
                                    if str(type(date_time)) == "<class 'datetime.datetime'>":
                                        pass
                                    else:
                                        print("in else")
                                        csv_success_flag = False
                                        remark += "Invalid Created Date, "
                                except Exception as e:
                                    print("In exception")
                                    logging.exception(f"{e}")
                                    csv_success_flag = False
                                    remark += "Invalid Created Date, "
                if self.fields[count] == FIELD_NAMES[5]:
                    try:
                        if '.' in col:
                            test_score = float(col)
                        else:
                            test_score = int(col)
                        if str(type(test_score)) == "<class 'float'>":
                            pass
                        elif str(type(test_score)) == "<class 'int'>":
                            pass
                        else:
                            remark += "Invalid Test Score, "
                    except Exception as e:
                        logging.exception(f"{e}")
                        remark += "Invalid Test Score, "
                if self.fields[count] == FIELD_NAMES[6]:
                    try:
                        if len(col) > 0:
                            pass
                        else:
                            remark += "incorrect resume link"
                    except Exception as e:
                        logging.exception(f"Error getting Resume Link {e}")
                        remark += "Invalid Resume Link, "
                if self.fields[count] == FIELD_NAMES[7]:
                    try:
                        if '.' in col:
                            notice_period = float(col)
                        else:
                            notice_period = int(col)
                        if str(type(notice_period)) == "<class 'float'>":
                            pass
                        elif str(type(notice_period)) == "<class 'int'>":
                            pass
                        else:
                            remark += "Invalid Notice Period, "
                    except Exception as e:
                        logging.exception(f"{e}")
                        remark += "Invalid Notice Period, "
                if self.fields[count] == FIELD_NAMES[8]:
                    try:
                        if '.' in col:
                            c_ctc = float(col)
                        else:
                            c_ctc = int(col)
                        if str(type(c_ctc)) == "<class 'float'>":
                            pass
                        elif str(type(c_ctc)) == "<class 'int'>":
                            pass
                        else:
                            remark += "Invalid Current CTC, "
                    except Exception as e:
                        logging.exception(f"{e}")
                        remark += "Invalid Current CTC, "
                if self.fields[count] == FIELD_NAMES[9]:
                    try:
                        if '.' in col:
                            e_ctc = float(col)
                        else:
                            e_ctc = int(col)
                        if str(type(e_ctc)) == "<class 'float'>":
                            pass
                        elif str(type(e_ctc)) == "<class 'int'>":
                            pass
                        else:
                            remark += "Invalid Excepted CTC, "
                    except Exception as e:
                        logging.exception(f"{e}")
                        remark += "Invalid Excepted CTC, "
                if self.fields[count] == FIELD_NAMES[10]:
                    try:
                        if len(col) > 0:
                            query = f""" SELECT {TABLE_SKILL_ID} FROM {TABLE_SKILL_NAME} 
                            WHERE {TABLE_SKILL_TITLE} = '{col}'"""
                            skill_result = read_query(self.connection, query)
                            if len(skill_result) > 0 or skill_result is not None:
                                skill_list.append({str(col).upper(): int(skill_result[0][0])})
                                skill1_id = int(skill_result[0][0])
                            else:
                                csv_success_flag = False
                                remark += "Invalid Skill1, "
                                skill1_id = None
                        else:
                            remark += "Invalid Skill1, "
                            logging.warning(f"Skill id for {col} does not found")
                            skill1_id = None
                    except Exception as e:
                        logging.exception(f"Error getting Skill1 {e}")
                        remark += "Invalid Skill1, "
                if self.fields[count] == FIELD_NAMES[12]:
                    try:
                        if '.' in col:
                            skill1_yr = float(col)
                        else:
                            skill1_yr = int(col)
                        if str(type(skill1_yr)) == "<class 'float'>":
                            pass
                        elif str(type(skill1_yr)) == "<class 'int'>":
                            pass
                        else:
                            remark += "Invalid Skill1 Years, "
                    except Exception as e:
                        logging.exception(f"{e}")
                        remark += "Invalid Skill1 Years, "
                if self.fields[count] == FIELD_NAMES[13]:
                    try:
                        if len(col) > 0:
                            query = f""" SELECT {TABLE_SKILL_ID} FROM {TABLE_SKILL_NAME} 
                            WHERE {TABLE_SKILL_TITLE} = '{col}'"""
                            skill_result = read_query(self.connection, query)
                            if len(skill_result) > 0 or skill_result is not None:
                                skill_list.append({str(col).upper(): int(skill_result[0][0])})
                                skill2_id = int(skill_result[0][0])
                            else:
                                csv_success_flag = False
                                remark += "Invalid Skill2, "
                                skill2_id = None
                        else:
                            remark += "Invalid Skill2, "
                            logging.warning(f"Skill id for {col} does not found")
                            skill2_id = None
                    except Exception as e:
                        logging.exception(f"Error getting Skill2 {e}")
                        remark += "Invalid Skill2, "
                if self.fields[count] == FIELD_NAMES[15]:
                    try:
                        if '.' in col:
                            skill2_yr = float(col)
                        else:
                            skill2_yr = int(col)
                        if str(type(skill2_yr)) == "<class 'float'>":
                            pass
                        elif str(type(skill2_yr)) == "<class 'int'>":
                            pass
                        else:
                            remark += "Invalid Skill2 Years, "
                    except Exception as e:
                        logging.exception(f"{e}")
                        remark += "Invalid Skill2 Years, "
                if self.fields[count] == FIELD_NAMES[16]:
                    try:
                        if len(col) > 0:
                            query = f""" SELECT {TABLE_SKILL_ID} FROM {TABLE_SKILL_NAME} 
                            WHERE {TABLE_SKILL_TITLE} = '{col}'"""
                            skill_result = read_query(self.connection, query)
                            if len(skill_result) > 0 or skill_result is not None:
                                skill_list.append({str(col).upper(): int(skill_result[0][0])})
                                skill3_id = int(skill_result[0][0])
                            else:
                                csv_success_flag = False
                                remark += "Invalid Skill3, "
                                skill3_id = None
                        else:
                            remark += "Invalid Skill3, "
                            logging.warning(f"Skill id for {col} does not found")
                            skill3_id = None
                    except Exception as e:
                        logging.exception(f"Error getting Skill3 {e}")
                        remark += "Invalid Skill3, "
                if self.fields[count] == FIELD_NAMES[18]:
                    try:
                        if '.' in col:
                            skill3_yr = float(col)
                        else:
                            skill3_yr = int(col)
                        if str(type(skill3_yr)) == "<class 'float'>":
                            pass
                        elif str(type(skill3_yr)) == "<class 'int'>":
                            pass
                        else:
                            remark += "Invalid Skill3 Years, "
                    except Exception as e:
                        logging.exception(f"incorrect skill3 years {e}")
                        remark += "Invalid Skill3 Years, "
                if count < len(self.fields) - 1:
                    count += 1

            if len(skill_list) > 0:
                try:
                    skill_count = 0
                    for skill in skill_list:
                        if skill in self.job_skills:
                            skill_count += 1
                    if skill_count > 0:
                        logging.info(f"{skill_count} skill match for given job")
                    else:
                        csv_success_flag = False
                        remark += "Given Skills does not match with required job skills"
                except Exception as e:
                    logging.exception(f"Error validating skill ids with job skill ids {e}")
                    csv_success_flag = False
                    remark += "Given Skills does not match with required job skills"
            else:
                logging.warning(f"All the three given skills are invalid")
                csv_success_flag = False
                remark += "Given Skills does not match with required job skills"

            row.insert(11, skill1_id)
            row.insert(14, skill2_id)
            row.insert(17, skill3_id)
            row.append(remark)
            if csv_success_flag:
                self.csv_success_row_list.append(row)
            else:
                self.csv_error_row_list.append(row)
        self.connection.close()

    # writing to csv file+
    def csv_file_writing(self):
        file_details = {}
        try:
            for i in range(0, 2):
                if i == 0:
                    file_name = self.csv_success_file_name
                    row_list = self.csv_success_row_list
                else:
                    file_name = self.csv_error_file_name
                    row_list = self.csv_error_row_list
                with open(file_name, 'w') as csvfile:
                    # creating a csv writer object
                    csvwriter = csv.writer(csvfile)

                    # writing the fields
                    csvwriter.writerow(FIELD_NAMES)

                    # writing the data rows
                    csvwriter.writerows(row_list)
                logging.info(f"{file_name} Written Successfully")

                with open(file_name, 'r') as csvfile:
                    # creating a csv reader object
                    csvreader = csv.reader(csvfile)

                    row_count = 0
                    for row in csvreader:
                        row_count += 1

                    file_details[file_name] = int(row_count) - 1

            return True, file_details
        except Exception as e:
            logging.exception(f"Error writing file: {e}")
            return False, file_details
########################################################################################################################


########################################################################################################################
class ExcelFileValidation:
    def __init__(self, path, job_skills):
        self.path = path
        self.sheet_obj = None
        self.row = 0
        self.column = 0
        self.excel_success_row_list = []
        self.excel_success_row_list.append(FIELD_NAMES)
        self.excel_error_row_list = []
        self.excel_error_row_list.append(FIELD_NAMES)
        self.excel_success_file_name = self.path.replace('.xlsx', '-Success.xlsx')
        self.excel_error_file_name = path.replace('.xlsx', '-Error.xlsx')
        self.job_skills = job_skills

        self.connection = create_server_connection()

        self.excel_file_reading()
        self.excel_file_validation()

    def excel_file_reading(self):
        try:
            # To open the workbook
            # workbook object is created
            wb_obj = openpyxl.load_workbook(self.path)

            # Get workbook active sheet object
            # from the active attribute
            self.sheet_obj = wb_obj.active

            # Getting the value of maximum rows
            # and column
            self.row = self.sheet_obj.max_row
            self.column = self.sheet_obj.max_column
        except Exception as e:
            logging.exception(f"Error reading {self.path} file {e}")

    def excel_file_validation(self):
        for j in range(2, self.row + 1):
            success_flag = True
            # time.sleep(2)
            row_list = []
            remark = ""
            skill_list = []
            for i in range(1, self.column + 1):
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[0]:
                    try:
                        first_name = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(first_name)
                        if first_name is None or type(first_name) != str:
                            success_flag = False
                            remark += "Invalid first name, "
                    except Exception as e:
                        logging.exception(f"Error reading first name {e}")
                        success_flag = False
                        remark += "Invalid first name, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[1]:
                    try:
                        last_name = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(last_name)
                        if last_name is None or type(last_name) != str:
                            success_flag = False
                            remark += "Invalid last name, "
                    except Exception as e:
                        logging.exception(f"Error reading last name {e}")
                        success_flag = False
                        remark += "Invalid last name, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[2]:
                    try:
                        mob_no = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(mob_no)
                        if mob_no is None or type(mob_no) != int or len(str(mob_no)) < 10:
                            success_flag = False
                            remark += "Invalid Mob No., "
                    except Exception as e:
                        logging.exception(f"Error reading MobNo. {e}")
                        success_flag = False
                        remark += "Invalid Mob No., "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[3]:
                    try:
                        email_id = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(email_id)
                        if email_id is None or email_validation(email_id) is False:
                            success_flag = False
                            remark += "Invalid Email-ID, "
                    except Exception as e:
                        logging.exception(f"Error reading Email Id {e}")
                        success_flag = False
                        remark += "Invalid Email-ID, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[4]:
                    try:
                        created_time = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(created_time)
                        if created_time is None or type(created_time) != datetime:
                            success_flag = False
                            remark += "Invalid Created Time, "
                    except Exception as e:
                        logging.exception(f"Error reading Created Time {e}")
                        success_flag = False
                        remark += "Invalid Created Time, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[5]:
                    try:
                        test_score = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(test_score)
                        if test_score is None or (type(test_score) != float and type(test_score) != int):
                            remark += "Invalid Test Score, "
                    except Exception as e:
                        logging.exception(f"Error reading Test Score {e}")
                        remark += "Invalid Test Score, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[6]:
                    try:
                        linked_resume = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(linked_resume)
                        if linked_resume is None or type(linked_resume) != str:
                            remark += "Invalid Resume Link, "
                    except Exception as e:
                        logging.exception(f"Error reading Resume Link {e}")
                        remark += "Invalid Resume Link, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[7]:
                    try:
                        notice_period = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(notice_period)
                        if notice_period is None or (type(notice_period) != float and type(notice_period) != int):
                            remark += "Invalid Notice Period, "
                    except Exception as e:
                        logging.exception(f"Error reading Notice Period {e}")
                        remark += "Invalid Notice Period, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[8]:
                    try:
                        current_ctc = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(current_ctc)
                        if current_ctc is None or (type(current_ctc) != float and type(current_ctc) != int):
                            remark += "Invalid Current CTC, "
                    except Exception as e:
                        logging.exception(f"Error reading Current CTC {e}")
                        remark += "Invalid Current CTC, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[9]:
                    try:
                        expected_ctc = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(expected_ctc)
                        if expected_ctc is None or (type(expected_ctc) != float and type(expected_ctc) != int):
                            remark += "Invalid Expected CTC, "
                    except Exception as e:
                        logging.exception(f"Error reading Expected CTC {e}")
                        remark += "Invalid Expected CTC, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[10]:
                    try:
                        skill1 = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(skill1)
                        if skill1 is None or type(skill1) != str:
                            remark += "Invalid Skill1, "
                            row_list.append(None)
                        else:
                            query = f""" SELECT {TABLE_SKILL_ID} FROM {TABLE_SKILL_NAME} 
                            WHERE {TABLE_SKILL_TITLE} = '{skill1}'"""
                            skill_result = read_query(self.connection, query)
                            if len(skill_result) > 0:
                                skill_list.append({str(skill1).upper(): int(skill_result[0][0])})
                                row_list.append(int(skill_result[0][0]))
                            else:
                                row_list.append(None)
                                logging.warning(f"Skill id for {skill1} does not found")
                                remark += "Invalid Skill1, "
                    except Exception as e:
                        logging.exception(f"Error reading Skill1 {e}")
                        remark += "Invalid Skill1, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[12]:
                    try:
                        skill1_years = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(skill1_years)
                        if skill1_years is None or (type(skill1_years) != float and type(skill1_years) != int):
                            remark += "Invalid Skill1 Years, "
                    except Exception as e:
                        logging.exception(f"Error reading Skill1 years {e}")
                        remark += "Invalid Skill1 Years, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[13]:
                    try:
                        skill2 = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(skill2)
                        if skill2 is None or type(skill2) != str:
                            remark += "Invalid Skill2, "
                            row_list.append(None)
                        else:
                            query = f""" SELECT {TABLE_SKILL_ID} FROM {TABLE_SKILL_NAME} 
                            WHERE {TABLE_SKILL_TITLE} = '{skill2}'"""
                            skill_result = read_query(self.connection, query)
                            if len(skill_result) > 0:
                                skill_list.append({str(skill2).upper(): int(skill_result[0][0])})
                                row_list.append(int(skill_result[0][0]))
                            else:
                                row_list.append(None)
                                logging.warning(f"Skill id for {skill2} does not found")
                                remark += "Invalid Skill2, "
                    except Exception as e:
                        logging.exception(f"Error reading Skill2 {e}")
                        remark += "Invalid Skill2, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[15]:
                    try:
                        skill2_years = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(skill2_years)
                        if skill2_years is None or (type(skill2_years) != float and type(skill2_years) != int):
                            remark += "Invalid Skill2 Years, "
                    except Exception as e:
                        logging.exception(f"Error reading Skill2 years {e}")
                        remark += "Invalid Skill2 Years, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[16]:
                    try:
                        skill3 = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(skill3)
                        if skill3 is None or type(skill3) != str:
                            remark += "Invalid Skill3, "
                            row_list.append(None)
                        else:
                            query = f""" SELECT {TABLE_SKILL_ID} FROM {TABLE_SKILL_NAME} 
                            WHERE {TABLE_SKILL_TITLE} = '{skill3}'"""
                            skill_result = read_query(self.connection, query)
                            if len(skill_result) > 0:
                                skill_list.append({str(skill3).upper(): int(skill_result[0][0])})
                                row_list.append(int(skill_result[0][0]))
                            else:
                                row_list.append(None)
                                logging.warning(f"Skill id for {skill3} does not found")
                                remark += "Invalid Skill3, "
                    except Exception as e:
                        logging.exception(f"Error reading Skill3 {e}")
                        remark += "Invalid Skill3, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[18]:
                    try:
                        skill3_years = self.sheet_obj.cell(row=j, column=i).value
                        row_list.append(skill3_years)
                        if skill3_years is None or (type(skill3_years) != float and type(skill3_years) != int):
                            remark += "Invalid Skill3 Years, "
                    except Exception as e:
                        logging.exception(f"Error reading Skill3 years {e}")
                        remark += "Invalid Skill3 Years, "
            if len(skill_list) > 0:
                try:
                    skill_count = 0
                    for skill in skill_list:
                        if skill in self.job_skills:
                            skill_count += 1
                    if skill_count > 0:
                        logging.info(f"{skill_count} skill match for given job")
                    else:
                        success_flag = False
                        remark += "Given Skills does not match with required job skills"
                except Exception as e:
                    logging.exception(f"Error validating skill ids with job skill ids {e}")
                    success_flag = False
                    remark += "Given Skills does not match with required job skills"

            row_list.append(remark)

            if success_flag:
                self.excel_success_row_list.append(row_list)
            else:
                self.excel_error_row_list.append(row_list)
        self.connection.close()

    def excel_file_writing(self):
        file_details = {}
        try:
            for i in range(0, 2):
                if i == 0:
                    file_name = self.excel_success_file_name
                    row_list = self.excel_success_row_list
                else:
                    file_name = self.excel_error_file_name
                    row_list = self.excel_error_row_list

                # Call a Workbook() function of openpyxl
                # to create a new blank Workbook object
                workbook = Workbook()

                # Anytime you modify the Workbook object
                # or its sheets and cells, the spreadsheet
                # file will not be saved until you call
                # the save() workbook method.

                workbook.save(filename=file_name)

                wb_success = openpyxl.load_workbook(file_name)

                sheet_success = wb_success.active

                for rows in tuple(row_list):
                    sheet_success.append(tuple(rows))

                wb_success.save(file_name)
                logging.info(f"{file_name} Written Successfully")

                # To open the workbook
                wb_obj = openpyxl.load_workbook(file_name)

                sheet_obj = wb_obj.active

                row = sheet_obj.max_row

                file_details[file_name] = row-1
            return True, file_details
        except Exception as e:
            logging.exception(f"Error in writing file: {e}")
            return False, file_details
########################################################################################################################


def reading_config_file():
    f = open(CONFIG_FILE_PATH, "r")
    data = json.load(f)
    global HOST_NAME
    HOST_NAME = data["host_name"]
    global USER_NAME
    USER_NAME = data["user_name"]
    global PASSWORD
    PASSWORD = data["password"]
    global DB_NAME
    DB_NAME = data["db_name"]


def create_server_connection():
    connect = None
    try:
        connect = mysql.connector.connect(
            host=HOST_NAME,
            user=USER_NAME,
            passwd=PASSWORD,
            database=DB_NAME
        )
        logging.info("Connection Successful")
    except Error as err:
        logging.exception(f"Error: {err}")
    return connect


def execute_query(connect, query):
    cursor = connect.cursor()
    try:
        cursor.execute(query)
        connect.commit()
        logging.info("Query Execute Successfully")
    except Error as err:
        logging.exception(f"Error : {err}")


def read_query(connect, query):
    cursor = connect.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()

    except Error as err:
        logging.exception(f"Error: {err}")
    return result


def email_validation(email_string):
    try:
        pat = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        if re.match(pat, email_string):
            return True
        else:
            return False
    except Exception as e:
        logging.exception(f"error in email validation {e}")
        return False


def thread_execution(row_id, path, job_id):
    try:
        logging.info(threading.current_thread().name)
        connection = create_server_connection()

        query = f'''select {TABLE_JOB_SKILL_SKILL_NAME}, {TABLE_JOB_SKILL_SKILL_ID} 
        From {TABLE_JOB_SKILL_NAME} where {TABLE_JOB_SKILL_JOB_ID} = {job_id};'''

        job_skill_result = read_query(connection, query)

        job_skill_required = []

        print(job_skill_result)
        if job_skill_result is not None and len(job_skill_result) > 0:
            for job_skill in job_skill_result:
                job_skill_required.append({str(job_skill[0]): int(job_skill[1])})
        else:
            logging.warning(f"Job Skills does not found for given Job ID: {job_id}")

        print(job_skill_required)

        file_writing_flag = False
        file_details = {}
        if ".csv" in path:
            csv_obj = CsvFileValidation(path, job_skill_required)
            file_writing_flag, file_details = csv_obj.csv_file_writing()
        elif ".xlsx" in path:
            excel_obj = ExcelFileValidation(path, job_skill_required)
            file_writing_flag, file_details = excel_obj.excel_file_writing()
        else:
            query = f'''UPDATE {TABLE_NAME}
                    SET {TABLE_COLUMN_PROCESS_STATE} = '-1'
                    WHERE {TABLE_COLUMN_ID} = '{row_id}';
                    '''
            execute_query(connection, query)
            logging.warning(f"Given File Path for row-id {row_id} is invalid")

        time_now = datetime.now()
        time_now = time_now.strftime('%Y-%m-%d %H:%M:%S')
        print(file_details)
        if file_writing_flag:
            query = f'''UPDATE {TABLE_NAME}
                        SET {TABLE_COLUMN_PROCESS_STATE} = '1'
                        WHERE {TABLE_COLUMN_ID} = '{row_id}';
                        '''
            execute_query(connection, query)

            success_file_path = ""
            success_file_rows = 0
            error_file_path = ""
            error_file_rows = 0

            for key in file_details:
                if "Success" in key:
                    success_file_path = key
                    success_file_rows = file_details[key]
                elif "Error" in key:
                    error_file_path = key
                    error_file_rows = file_details[key]
                else:
                    success_file_path = ""
                    success_file_rows = 0
                    error_file_path = ""
                    error_file_rows = 0

            print(success_file_path)
            print(success_file_rows)
            print(error_file_path)
            print(error_file_rows)

            query = f'''UPDATE {TABLE_NAME}
                        SET {TABLE_COLUMN_FILE_UPDATE_TIME} = '{time_now}',
                        {TABLE_COLUMN_SUCCESS_FILE_PATH} = '{success_file_path}',
                        {TABLE_COLUMN_ERROR_FILE_PATH} = '{error_file_path}',
                        {TABLE_COLUMN_SUCCESS_FILE_ROWS} = {int(success_file_rows)},
                        {TABLE_COLUMN_ERROR_FILE_ROWS} = {int(error_file_rows)}
                        WHERE {TABLE_COLUMN_ID} = '{row_id}';
                        '''
            execute_query(connection, query)
        else:
            logging.info("No file to Update")
        connection.close()
    except Exception as e:
        logging.exception(f"Error in Tread Execution Method {e}")


def main():
    reading_config_file()
    while True:
        try:
            connection = create_server_connection()

            query = f'''
            select {TABLE_COLUMN_ID}, {TABLE_COLUMN_PATH}, {TABLE_COLUMN_JOB_ID} From 
            {TABLE_NAME} where {TABLE_COLUMN_PROCESS_STATE} = '0';
            '''

            query_results = read_query(connection, query)

            if query_results is not None and len(query_results) > 0:
                for result in query_results:
                    row_id = int(result[0])
                    path = str(result[1])
                    job_id = int(result[2])

                    t1 = threading.Thread(target=thread_execution, name='t1', args=(row_id, path, job_id))
                    t2 = threading.Thread(target=thread_execution, name='t2', args=(row_id, path, job_id))

                    logging.info(threading.current_thread().ident)

                    if row_id % 2 == 1:
                        t1.start()
                    if row_id % 2 == 0:
                        t2.start()

            else:
                logging.info("No new file to update")
            connection.close()
        except Exception as e:
            logging.exception(f"error in main method {e}")
        print("Sleeping for 30 sec")
        logging.info("Sleeping for 30 sec")
        time.sleep(30)


if __name__ == "__main__":
    main()
