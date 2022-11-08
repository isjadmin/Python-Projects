# Python program to read an excel file

# import openpyxl module
import openpyxl
from openpyxl import Workbook

# Give the location of the file
path = "Test-Book1.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows:", row)
print("Total Columns:", column)

# printing the value of first column
# Loop will print all values
# of first column
print("\nValue of first column")
for i in range(1, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print(cell_obj.value)

# printing the value of first column
# Loop will print all values
# of first row
print("\nValue of first row")
for i in range(1, column + 1):
    cell_obj = sheet_obj.cell(row=2, column=i)
    print(cell_obj.value, end=" ")

workbook = Workbook()

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
workbook.save(filename="sample.xlsx")
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute
sheet = wb.active
data = []

for i in range(1, row+1):
    for j in range(1, column + 1):
        cell_obj = sheet_obj.cell(row=i, column=j)
        print(cell_obj.value)
        if type(cell_obj.value) is not int and 'Birth' in cell_obj.value:
            if 'Birthdate' in cell_obj.value:
                birthdate = sheet_obj.cell(row=i, column=j+1)
            if 'Birthmonth' in cell_obj.value:
                birthmonth = sheet_obj.cell(row=i, column=j+1)
            if 'Birthyear' in cell_obj.value:
                birthyear = sheet_obj.cell(row=i, column=j+1)
                birthday = str(birthdate.value) + '-' + str(birthmonth.value) + '-' + str(birthyear.value)
            if 'Birthtime in hr' in cell_obj.value:
                birthtime_hr = sheet_obj.cell(row=i, column=j+1)
            if 'Birthtime in min' in cell_obj.value:
                birthtime_min = sheet_obj.cell(row=i, column=j+1)
                birthtime = str(birthtime_hr.value) + ':' + str(birthtime_min.value)
                print(f'Birthday : {birthday}, Birthtime : {birthtime}')
        data[i-1][j-1] = cell_obj.value

print(data)
