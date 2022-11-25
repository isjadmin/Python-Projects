import openpyxl
from openpyxl import Workbook
import re
from datetime import datetime
import mysql.connector
from mysql.connector import Error
import csv
import logging


FIELD_NAMES = ['name', 'mob no', 'email', 'created time', 'testscore', 'linked resume', 'notice period',
               'current ctc', 'expected ctc', 'skill1', 'skill1 ID', 'skill1 years', 'skill2', 'skill2 ID',
               'skill2 years', 'skill3', 'skill3 ID', 'skill3 years', 'remark']

HOST_NAME = "localhost"
USER_NAME = "root"
PASSWORD = "Smprajkta4$"
DB_NAME = "jobztop"

TABLE_NAME = "upload"
TABLE_COLUMN_ID = "id"
TABLE_COLUMN_PATH = "path"
TABLE_COLUMN_PROCESS_STATE = "processed_state"

CANDIDATE_TABLE_NAME = "candidate"
CANDIDATE_TABLE_FIRST_NAME = "FirstName"
CANDIDATE_TABLE_COLUMN_ID = "Id"
CANDIDATE_TABLE_COLUMN_MOB_NO = "Mobile"
CANDIDATE_TABLE_COLUMN_CCTC = "CTC"
CANDIDATE_TABLE_COLUMN_ECTC = "ECTC"
CANDIDATE_TABLE_COLUMN_EMAIL_ID = "EmailId"
CANDIDATE_TABLE_COLUMN_NOTICE_PERIOD = "NoticePeriod"

CANDIDATE_SKILL_TABLE_NAME = "candidateskill"
CANDIDATE_SKILL_TABLE_CANDIDATE_ID = "CandidateId"
CANDIDATE_SKILL_TABLE_SKILL_ID = "SkillId"


########################################################################################################################
class CsvFileValidation:
    def __init__(self, path):
        self.path = path
        self.rows = []
        self.fields = []

        self.csv_file_reading()

    # reading csv file
    def csv_file_reading(self):
        with open(self.path, 'r') as csvfile:
            # creating a csv reader object
            csvreader = csv.reader(csvfile)

            # extracting field names through first row
            self.fields = next(csvreader)

            # extracting each data row one by one
            for row in csvreader:
                self.rows.append(row)

            # get total number of rows
            logging.info("Total no. of rows: %d" % csvreader.line_num)

    def csv_required_details(self):
        candidate_details = []
        candidate_skill_id = []
        for row in self.rows:
            # parsing each column of a row
            count = 0
            candidate_detail_list = []
            skill_id_list = []
            if len(row) == 0:
                continue
            for col in row:
                if self.fields[count] == FIELD_NAMES[0]:
                    if len(col) > 0:
                        candidate_detail_list.append(col)
                    else:
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[1]:
                    if len(col) > 0:
                        candidate_detail_list.append(int(col))
                        skill_id_list.append(int(col))
                    else:
                        candidate_detail_list.append('0')
                        skill_id_list.append('0')
                if self.fields[count] == FIELD_NAMES[2]:
                    if len(col) > 0:
                        print(col)
                        candidate_detail_list.append(col)
                    else:
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[6]:
                    if len(col) > 0:
                        print(col)
                        candidate_detail_list.append(int(col))
                    else:
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[7]:
                    if len(col) > 0:
                        print(col)
                        candidate_detail_list.append(float(col))
                    else:
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[8]:
                    if len(col) > 0:
                        candidate_detail_list.append(float(col))
                    else:
                        candidate_detail_list.append('0')
                if self.fields[count] == FIELD_NAMES[10]:
                    if len(col) > 0:
                        print(col)
                        skill_id_list.append(int(col))
                    else:
                        skill_id_list.append('0')
                if self.fields[count] == FIELD_NAMES[13]:
                    if len(col) > 0:
                        skill_id_list.append(int(col))
                    else:
                        skill_id_list.append('0')
                if self.fields[count] == FIELD_NAMES[16]:
                    if len(col) > 0:
                        skill_id_list.append(int(col))
                    else:
                        skill_id_list.append('0')
                if count < len(self.fields) - 1:
                    count += 1

            candidate_details.append(candidate_detail_list)
            candidate_skill_id.append(skill_id_list)

        return candidate_details, candidate_skill_id

########################################################################################################################


########################################################################################################################
class ExcelFileValidation:
    def __init__(self, path):
        self.path = path
        self.sheet_obj = None
        self.row = 0
        self.column = 0
        self.excel_success_row_list = []
        self.excel_success_row_list.append(FIELD_NAMES)
        self.excel_error_row_list = []
        self.excel_error_row_list.append(FIELD_NAMES)
        self.excel_success_file_name = self.path.removesuffix('.xlsx') + '-Success.xlsx'
        self.excel_error_file_name = path.removesuffix('.xlsx') + '-Error.xlsx'

        self.connection = create_server_connection()

        self.excel_file_reading()

    def excel_file_reading(self):
        # To open the workbook
        # workbook object is created
        wb_obj = openpyxl.load_workbook(self.path)

        # Get workbook active sheet object
        # from the active attribute
        self.sheet_obj = wb_obj.active

        # Getting the value of maximum rows
        # and column
        self.row = self.sheet_obj.max_row
        self.column = self.sheet_obj.max_column

    def excel_required_details(self):
        candidate_details = []
        candidate_skill_id = []
        for j in range(2, self.row + 1):
            candidate_detail_list = []
            skill_id_list = []
            for i in range(1, self.column + 1):
                # print(sheet_obj.cell(row=j, column=i).value)
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[0]:
                    # print('name: ', type(sheet_obj.cell(row=j, column=i).value))
                    name = self.sheet_obj.cell(row=j, column=i).value
                    if name is None or type(name) != str:
                        candidate_detail_list.append('0')
                    else:
                        candidate_detail_list.append(name)
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[1]:
                    # print('mob no: ', type(sheet_obj.cell(row=j, column=i).value))
                    mob_no = self.sheet_obj.cell(row=j, column=i).value
                    if mob_no is None or type(mob_no) != int or len(str(mob_no)) < 10:
                        candidate_detail_list.append('0')
                        skill_id_list.append('0')
                    else:
                        candidate_detail_list.append(mob_no)
                        skill_id_list.append(mob_no)
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[2]:
                    # print('email: ', type(sheet_obj.cell(row=j, column=i).value))
                    email_id = self.sheet_obj.cell(row=j, column=i).value
                    if email_id is None:
                        candidate_detail_list.append('0')
                    else:
                        candidate_detail_list.append(email_id)
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[6]:
                    # print('notice period: ', type(sheet_obj.cell(row=j, column=i).value))
                    notice_period = self.sheet_obj.cell(row=j, column=i).value
                    if notice_period is None or (type(notice_period) != float and type(notice_period) != int):
                        candidate_detail_list.append('0')
                    else:
                        candidate_detail_list.append(notice_period)
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[7]:
                    # print('current ctc: ', type(sheet_obj.cell(row=j, column=i).value))
                    current_ctc = self.sheet_obj.cell(row=j, column=i).value
                    if current_ctc is None or (type(current_ctc) != float and type(current_ctc) != int):
                        candidate_detail_list.append('0')
                    else:
                        candidate_detail_list.append(current_ctc)
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[8]:
                    # print('expected ctc: ', type(sheet_obj.cell(row=j, column=i).value))
                    expected_ctc = self.sheet_obj.cell(row=j, column=i).value
                    if expected_ctc is None or (type(expected_ctc) != float and type(expected_ctc) != int):
                        candidate_detail_list.append('0')
                    else:
                        candidate_detail_list.append(expected_ctc)
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[10]:
                    skill1_id = self.sheet_obj.cell(row=j, column=i).value
                    if skill1_id is None:
                        skill_id_list.append('0')
                    else:
                        skill_id_list.append(skill1_id)
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[13]:
                    skill2_id = self.sheet_obj.cell(row=j, column=i).value
                    if skill2_id is None:
                        skill_id_list.append('0')
                    else:
                        skill_id_list.append(skill2_id)
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[16]:
                    skill3_id = self.sheet_obj.cell(row=j, column=i).value
                    if skill3_id is None:
                        skill_id_list.append('0')
                    else:
                        skill_id_list.append(skill3_id)

            candidate_details.append(candidate_detail_list)
            candidate_skill_id.append(skill_id_list)

        return candidate_details, candidate_skill_id

########################################################################################################################


def create_server_connection():
    connect = None
    try:
        connect = mysql.connector.connect(
            host=HOST_NAME,
            user=USER_NAME,
            passwd=PASSWORD,
            database=DB_NAME
        )
        print("Connection Successful")
    except Error as err:
        print(f"Error: {err}")
    return connect


def execute_query(connect, query):
    cursor = connect.cursor()
    try:
        cursor.execute(query)
        connect.commit()
        print("Query Execute Successfully")
    except Error as err:
        print(f"Error : {err}")


def read_query(connect, query):
    cursor = connect.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()

    except Error as err:
        print(f"Error: {err}")
    return result


def update_candidate_table(candidate_details):
    connection = create_server_connection()
    for detail in candidate_details:
        query = f'''select {CANDIDATE_TABLE_COLUMN_ID} From {CANDIDATE_TABLE_NAME} 
        where {CANDIDATE_TABLE_COLUMN_MOB_NO} = {detail[1]}'''

        print(detail[2])

        detail_result = read_query(connection, query)

        if detail_result is not None and len(detail_result) > 0:
            update_query = f'''UPDATE {CANDIDATE_TABLE_NAME} SET `{CANDIDATE_TABLE_COLUMN_CCTC}` = {detail[4]}, 
            `{CANDIDATE_TABLE_COLUMN_ECTC}` = {detail[5]}, `{CANDIDATE_TABLE_COLUMN_EMAIL_ID}` = {detail[2]}, 
            `{CANDIDATE_TABLE_FIRST_NAME}` = {detail[0]}, `{CANDIDATE_TABLE_COLUMN_NOTICE_PERIOD}` = {detail[3]} 
            WHERE ({CANDIDATE_TABLE_COLUMN_ID} = {detail_result[0]});'''

            execute_query(connection, update_query)

        else:
            insert_query = f'''INSERT INTO {CANDIDATE_TABLE_NAME} (`{CANDIDATE_TABLE_COLUMN_CCTC}`, 
            `{CANDIDATE_TABLE_COLUMN_ECTC}`, `{CANDIDATE_TABLE_COLUMN_EMAIL_ID}`, `{CANDIDATE_TABLE_FIRST_NAME}`, 
            `{CANDIDATE_TABLE_COLUMN_MOB_NO}`, `{CANDIDATE_TABLE_COLUMN_NOTICE_PERIOD}`) VALUES ({detail[4]}, 
            {detail[5]}, '{detail[2]}', '{detail[0]}', {detail[1]}, {detail[3]});'''

            execute_query(connection, insert_query)

    connection.close()


def candidate_skill_table_update(candidate_skills):
    connection = create_server_connection()
    for candidate_skill in candidate_skills:
        id_query = f'''select {CANDIDATE_TABLE_COLUMN_ID} From {CANDIDATE_TABLE_NAME} 
        where {CANDIDATE_TABLE_COLUMN_MOB_NO} = {candidate_skill[0]}'''
        id_result = read_query(connection, id_query)

        if id_result is not None and len(id_result) > 0:
            print(id_result[0])
            for skill in candidate_skill[1:]:
                skill_check_query = f'''select {CANDIDATE_SKILL_TABLE_SKILL_ID} from {CANDIDATE_SKILL_TABLE_NAME} 
                where {CANDIDATE_SKILL_TABLE_CANDIDATE_ID} = {skill}'''

                skill_check_result = read_query(connection, skill_check_query)

                if skill_check_result is not None and len(skill_check_result) > 0:
                    logging.info(f"Skill id {skill} already present for candidate id {int(id_result[0][0])}")
                    print(f"Skill id {skill} already present for candidate id {int(id_result[0][0])}")
                else:
                    skill_insert_query = f'''INSERT INTO {CANDIDATE_SKILL_TABLE_NAME} 
                    (`{CANDIDATE_SKILL_TABLE_CANDIDATE_ID}`, `{CANDIDATE_SKILL_TABLE_SKILL_ID}`) 
                    VALUES ({int(id_result[0][0])}, {skill});'''

                    execute_query(connection, skill_insert_query)
        else:
            logging.warning("Invalid candidate Id")


def main():
    try:
        connection = create_server_connection()

        query = f'''
        select {TABLE_COLUMN_ID}, {TABLE_COLUMN_PATH} From {TABLE_NAME} where {TABLE_COLUMN_PROCESS_STATE} = 1;
        '''

        query_results = read_query(connection, query)
        print(query_results)

        if query_results is not None and len(query_results) > 0:
            for result in query_results:
                # print(result)
                # print(type(result[1]))
                row_id = result[0]
                path = str(result[1])
                candidate_details = []
                candidate_skill_id = []
                if ".csv" in path:
                    path = path.removesuffix('.csv') + '-Success.csv'
                    csv_obj = CsvFileValidation(path)
                    candidate_details, candidate_skill_id = csv_obj.csv_required_details()
                    print(f"Candidate details from {path}: \n {candidate_details}")
                    print(f"Candidate skill ids from {path}: \n {candidate_skill_id}")
                elif ".xlsx" in path:
                    path = path.removesuffix('.xlsx') + '-Success.xlsx'
                    print(path)
                    excel_obj = ExcelFileValidation(path)
                    candidate_details, candidate_skill_id = excel_obj.excel_required_details()
                    print(f"Candidate details from {path}: \n {candidate_details}")
                    print(f"Candidate skill ids from {path}: \n {candidate_skill_id}")
                else:
                    logging.warning(f"Given File Path for row-id {row_id} is invalid")
                if len(candidate_details) > 0 and len(candidate_skill_id) > 0:
                    update_candidate_table(candidate_details)
                    candidate_skill_table_update(candidate_skill_id)
        else:
            logging.info("No new file to update")
        connection.close()
    except Exception as e:
        logging.exception(f"error in main method {e}")


if __name__ == "__main__":
    main()