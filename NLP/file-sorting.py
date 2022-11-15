import openpyxl
from openpyxl import Workbook
import re
from datetime import datetime
import time
import mysql.connector
from mysql.connector import Error
import csv
import logging


FIELD_NAMES = ['name', 'mob no', 'email', 'created time', 'testscore', 'linked resume', 'notice period',
               'current ctc', 'expected ctc', 'skill1', 'skill1 years', 'skill2', 'skill2 years',
               'skill3', 'skill3 years', 'remark']

HOST_NAME = "localhost"
USER_NAME = "root"
PASSWORD = "Smprajkta4$"
DB_NAME = "job_seeker_data"
TABLE_NAME = "Upload"
TABLE_COLUMN_ID = "id"
TABLE_COLUMN_PATH = "path"
TABLE_COLUMN_PROCESS_STATE = "processed_state"
TABLE_COLUMN_FILE_UPDATE_TIME = "file_update_time"


########################################################################################################################
class CsvFileValidation:
    def __init__(self, path):
        self.path = path
        self.rows = []
        self.fields = []
        self.csv_success_row_list = []
        self.csv_error_row_list = []
        self.csv_success_file_name = self.path.removesuffix('.csv') + '-Success.csv'
        self.csv_error_file_name = self.path.removesuffix('.csv') + '-Error.csv'

        self.csv_file_reading()
        self.csv_file_validation()

    # reading csv file
    def csv_file_reading(self):
        with open(self.path, 'r') as csvfile:
            # creating a csv reader object
            csvreader = csv.reader(csvfile)

            # extracting field names through first row
            self.fields = next(csvreader)

            # extracting each data row one by one
            for row in csvreader:
                self.rows.append(row)

            # get total number of rows
            logging.info("Total no. of rows: %d" % csvreader.line_num)

    def csv_file_validation(self):
        for row in self.rows:
            # parsing each column of a row
            count = 0
            csv_success_flag = True
            remark = ""
            if len(row) == 0:
                continue
            for col in row:
                # print(fields[count])
                '''print("%8s" % type(col), end="  ")
                print("%8s" % col, end="  ")'''
                if self.fields[count] == FIELD_NAMES[0]:
                    if len(col) > 0:
                        print(col)
                    else:
                        print("incorrect name")
                        csv_success_flag = False
                        remark += "Invalid Name, "
                if self.fields[count] == FIELD_NAMES[1]:
                    try:
                        print(int(col))
                    except Exception as e:
                        logging.exception(f"{e}")
                        print("incorrect mob no")
                        csv_success_flag = False
                        remark += "Invalid Mobile No., "
                if self.fields[count] == FIELD_NAMES[2]:
                    if email_validation(col):
                        print(col)
                    else:
                        csv_success_flag = False
                        print("incorrect email")
                        remark += "Invalid Email-ID, "
                if self.fields[count] == FIELD_NAMES[3]:
                    try:
                        date_time = col.split(',')
                        for i in range(len(date_time)):
                            date_time[i] = int(date_time[i])
                        print(date_time)
                        print(datetime.strftime(datetime(date_time[0], date_time[1], date_time[2], date_time[3],
                                                         date_time[4], date_time[5]),
                                                '%Y, %m, %d, %H, %M, %S'))
                    except Exception as e:
                        logging.exception(f"{e}")
                        csv_success_flag = False
                        print("Incorrect Created Date")
                        remark += "Invalid Created Date, "
                if self.fields[count] == FIELD_NAMES[4]:
                    try:
                        if '.' in col:
                            print(float(col))
                        else:
                            print(int(col))
                    except Exception as e:
                        logging.exception(f"{e}")
                        print("incorrect test score")
                        remark += "Invalid Test Score, "
                if self.fields[count] == FIELD_NAMES[5]:
                    if len(col) > 0:
                        print(col)
                    else:
                        print("incorrect linked resume")
                if self.fields[count] == FIELD_NAMES[6]:
                    try:
                        if '.' in col:
                            print(float(col))
                        else:
                            print(int(col))
                    except Exception as e:
                        logging.exception(f"{e}")
                        print("incorrect notice period")
                        remark += "Invalid Notice Period, "
                if self.fields[count] == FIELD_NAMES[7]:
                    try:
                        if '.' in col:
                            print(float(col))
                        else:
                            print(int(col))
                    except Exception as e:
                        logging.exception(f"{e}")
                        print("incorrect current ctc")
                        remark += "Invalid Current CTC, "
                if self.fields[count] == FIELD_NAMES[8]:
                    try:
                        if '.' in col:
                            print(float(col))
                        else:
                            print(int(col))
                    except Exception as e:
                        logging.exception(f"{e}")
                        print("incorrect expected ctc")
                        remark += "Invalid Excepted CTC, "
                if self.fields[count] == FIELD_NAMES[9]:
                    if len(col) > 0:
                        print(col)
                    else:
                        print("incorrect skill1")
                        remark += "Invalid Skill1, "
                if self.fields[count] == FIELD_NAMES[10]:
                    try:
                        if '.' in col:
                            print(float(col))
                        else:
                            print(int(col))
                    except Exception as e:
                        logging.exception(f"{e}")
                        print("incorrect skill1 years")
                        remark += "Invalid Skill1 Years, "
                if self.fields[count] == FIELD_NAMES[11]:
                    if len(col) > 0:
                        print(col)
                    else:
                        print("incorrect skill2")
                        remark += "Invalid Skill2, "
                if self.fields[count] == FIELD_NAMES[12]:
                    try:
                        if '.' in col:
                            print(float(col))
                        else:
                            print(int(col))
                    except Exception as e:
                        logging.exception(f"{e}")
                        print("incorrect skill2 years")
                        remark += "Invalid Skill2 Years, "
                if self.fields[count] == FIELD_NAMES[13]:
                    if len(col) > 0:
                        print(col)
                    else:
                        print("incorrect skill3")
                        remark += "Invalid Skill3, "
                if self.fields[count] == FIELD_NAMES[14]:
                    try:
                        if '.' in col:
                            print(float(col))
                        else:
                            print(int(col))
                    except Exception as e:
                        logging.exception(f"incorrect skill3 years {e}")
                        remark += "Invalid Skill3 Years, "
                if count < len(self.fields) - 1:
                    count += 1
            row.append(remark.removesuffix(', '))
            if csv_success_flag:
                self.csv_success_row_list.append(row)
            else:
                self.csv_error_row_list.append(row)

    # writing to csv file+
    def csv_file_writing(self):
        for i in range(0, 2):
            if i == 0:
                file_name = self.csv_success_file_name
                row_list = self.csv_success_row_list
            else:
                file_name = self.csv_error_file_name
                row_list = self.csv_error_row_list
            with open(file_name, 'w') as csvfile:
                # creating a csv writer object
                csvwriter = csv.writer(csvfile)

                # writing the fields
                csvwriter.writerow(FIELD_NAMES)

                # writing the data rows
                csvwriter.writerows(row_list)
        return True
########################################################################################################################


########################################################################################################################
class ExcelFileValidation:
    def __init__(self, path):
        self.path = path
        self.sheet_obj = None
        self.row = 0
        self.column = 0
        self.excel_success_row_list = []
        self.excel_success_row_list.append(FIELD_NAMES)
        self.excel_error_row_list = []
        self.excel_error_row_list.append(FIELD_NAMES)
        self.excel_success_file_name = self.path.removesuffix('.xlsx') + '-Success.xlsx'
        self.excel_error_file_name = path.removesuffix('.xlsx') + '-Error.xlsx'

        self.excel_file_reading()
        self.excel_file_validation()

    def excel_file_reading(self):
        # To open the workbook
        # workbook object is created
        wb_obj = openpyxl.load_workbook(self.path)

        # Get workbook active sheet object
        # from the active attribute
        self.sheet_obj = wb_obj.active

        # Getting the value of maximum rows
        # and column
        self.row = self.sheet_obj.max_row
        self.column = self.sheet_obj.max_column

    def excel_file_validation(self):
        for j in range(2, self.row + 1):
            success_flag = True
            # time.sleep(2)
            row_list = []
            remark = ""
            for i in range(1, self.column + 1):
                # print(sheet_obj.cell(row=j, column=i).value)
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[0]:
                    # print('name: ', type(sheet_obj.cell(row=j, column=i).value))
                    name = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(name)
                    if name is None or type(name) != str:
                        success_flag = False
                        remark += "Invalid name, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[1]:
                    # print('mob no: ', type(sheet_obj.cell(row=j, column=i).value))
                    mob_no = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(mob_no)
                    if mob_no is None or type(mob_no) != int or len(str(mob_no)) < 10:
                        success_flag = False
                        remark += "Invalid Mob No., "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[2]:
                    # print('email: ', type(sheet_obj.cell(row=j, column=i).value))
                    email_id = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(email_id)
                    if email_id is None or email_validation(email_id) is False:
                        success_flag = False
                        remark += "Invalid Email-ID, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[3]:
                    # print('created time: ', type(sheet_obj.cell(row=j, column=i).value))
                    created_time = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(created_time)
                    if created_time is None or type(created_time) != datetime:
                        success_flag = False
                        remark += "Invalid Created Time, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[4]:
                    # print('testscore: ', type(sheet_obj.cell(row=j, column=i).value))
                    test_score = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(test_score)
                    if test_score is None or (type(test_score) != float and type(test_score) != int):
                        remark += "Invalid Test Score, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[5]:
                    # print('linked resume: ', type(sheet_obj.cell(row=j, column=i).value))
                    linked_resume = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(linked_resume)
                    if linked_resume is None or type(linked_resume) != str:
                        remark += "Invalid Linked Resume, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[6]:
                    # print('notice period: ', type(sheet_obj.cell(row=j, column=i).value))
                    notice_period = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(notice_period)
                    if notice_period is None or (type(notice_period) != float and type(notice_period) != int):
                        remark += "Invalid Notice Period, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[7]:
                    # print('current ctc: ', type(sheet_obj.cell(row=j, column=i).value))
                    current_ctc = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(current_ctc)
                    if current_ctc is None or (type(current_ctc) != float and type(current_ctc) != int):
                        remark += "Invalid Current CTC, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[8]:
                    # print('expected ctc: ', type(sheet_obj.cell(row=j, column=i).value))
                    expected_ctc = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(expected_ctc)
                    if expected_ctc is None or (type(expected_ctc) != float and type(expected_ctc) != int):
                        remark += "Invalid Expected CTC, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[9]:
                    # print('skill1: ', type(sheet_obj.cell(row=j, column=i).value))
                    skill1 = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(skill1)
                    if skill1 is None or type(skill1) != str:
                        remark += "Invalid Skill1, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[10]:
                    # print('skill1 years: ', type(sheet_obj.cell(row=j, column=i).value))
                    skill1_years = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(skill1_years)
                    if skill1_years is None or (type(skill1_years) != float and type(skill1_years) != int):
                        remark += "Invalid Skill1 Years, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[11]:
                    # print('skill2: ', type(sheet_obj.cell(row=j, column=i).value))
                    skill2 = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(skill2)
                    if skill2 is None or type(skill2) != str:
                        remark += "Invalid Skill2, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[12]:
                    # print('skill2 years: ', type(sheet_obj.cell(row=j, column=i).value))
                    skill2_years = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(skill2_years)
                    if skill2_years is None or (type(skill2_years) != float and type(skill2_years) != int):
                        remark += "Invalid Skill2 Years, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[13]:
                    # print('skill 3: ', type(sheet_obj.cell(row=j, column=i).value))
                    skill3 = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(skill3)
                    if skill3 is None or type(skill3) != str:
                        remark += "Invalid Skill3, "
                if self.sheet_obj.cell(row=1, column=i).value == FIELD_NAMES[14]:
                    # print('skill3 years: ', type(sheet_obj.cell(row=j, column=i).value))
                    skill3_years = self.sheet_obj.cell(row=j, column=i).value
                    row_list.append(skill3_years)
                    if skill3_years is None or (type(skill3_years) != float and type(skill3_years) != int):
                        remark += "Invalid Skill3 Years, "
            row_list.append(remark.removesuffix(', '))
            print(row_list)
            print('--------------------------------------------------------------------------------------------------')
            if success_flag:
                print(success_flag)
                print(row_list)
                self.excel_success_row_list.append(row_list)
            else:
                print(success_flag)
                print(row_list)
                self.excel_error_row_list.append(row_list)

    def excel_file_writing(self):
        for i in range(0, 2):
            if i == 0:
                file_name = self.excel_success_file_name
                row_list = self.excel_success_row_list
            else:
                file_name = self.excel_error_file_name
                row_list = self.excel_error_row_list

            # Call a Workbook() function of openpyxl
            # to create a new blank Workbook object
            workbook = Workbook()

            # Anytime you modify the Workbook object
            # or its sheets and cells, the spreadsheet
            # file will not be saved until you call
            # the save() workbook method.

            workbook.save(filename=file_name)

            wb_success = openpyxl.load_workbook(file_name)

            sheet_success = wb_success.active

            for rows in tuple(row_list):
                sheet_success.append(tuple(rows))

            wb_success.save(file_name)
        return True
########################################################################################################################


def create_server_connection():
    connect = None
    try:
        connect = mysql.connector.connect(
            host=HOST_NAME,
            user=USER_NAME,
            passwd=PASSWORD,
            database=DB_NAME
        )
        print("Connection Successful")
    except Error as err:
        print(f"Error: {err}")
    return connect


def execute_query(connect, query):
    cursor = connect.cursor()
    try:
        cursor.execute(query)
        connect.commit()
        print("Query Execute Successfully")
    except Error as err:
        print(f"Error : {err}")


def read_query(connect, query):
    cursor = connect.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()

    except Error as err:
        print(f"Error: {err}")
    return result


def email_validation(email_string):
    try:
        pat = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        if re.match(pat, email_string):
            return True
        else:
            return False
    except Exception as e:
        logging.exception(f"error in email validation {e}")
        return False


def main():
    try:
        connection = create_server_connection()

        query = f'''
        select {TABLE_COLUMN_ID}, {TABLE_COLUMN_PATH} From {TABLE_NAME} where {TABLE_COLUMN_PROCESS_STATE} = 0;
        '''

        query_results = read_query(connection, query)
        print(query_results)

        if query_results is not None and len(query_results) > 0:
            for result in query_results:
                print(result)
                print(type(result[1]))
                row_id = result[0]
                path = str(result[1])
                file_writing_flag = False
                if ".csv" in path:
                    csv_obj = CsvFileValidation(path)
                    file_writing_flag = csv_obj.csv_file_writing()
                elif ".xlsx" in path:
                    excel_obj = ExcelFileValidation(path)
                    file_writing_flag = excel_obj.excel_file_writing()
                else:
                    logging.warning(f"Given File Path for row-id {row_id} is invalid")

                time_now = datetime.now()
                time_now = time_now.strftime('%Y-%m-%d %H:%M:%S')
                if file_writing_flag:
                    query = f'''
                    UPDATE {TABLE_NAME}
                    SET {TABLE_COLUMN_PROCESS_STATE} = 1
                    WHERE {TABLE_COLUMN_ID} = '{row_id}';
                    '''
                    execute_query(connection, query)

                    query = f'''
                    UPDATE {TABLE_NAME}
                    SET {TABLE_COLUMN_FILE_UPDATE_TIME} = '{time_now}'
                    WHERE {TABLE_COLUMN_ID} = '{row_id}';
                    '''
                    execute_query(connection, query)
                else:
                    logging.info("No file to Update")
        else:
            logging.info("No new file to update")
        connection.close()
    except Exception as e:
        logging.exception(f"error in main method {e}")


if __name__ == "__main__":
    main()
