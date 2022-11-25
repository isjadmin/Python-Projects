import openpyxl
from openpyxl import Workbook
import re
import datetime
from datetime import datetime
import time
import mysql.connector
from mysql.connector import Error

success_row_list = [['name', 'mob no', 'email', 'created time', 'testscore', 'linked resume', 'notice period',
                     'current ctc', 'expected ctc', 'skill1', 'skill1 ID', 'skill1 years', 'skill2', 'skill2 ID',
                     'skill2 years', 'skill3', 'skill3 ID', 'skill3 years', 'remark']]
error_row_list = [['name', 'mob no', 'email', 'created time', 'testscore', 'linked resume', 'notice period',
                   'current ctc', 'expected ctc', 'skill1', 'skill1 ID', 'skill1 years', 'skill2', 'skill2 ID',
                   'skill2 years', 'skill3', 'skill3 ID', 'skill3 years', 'remark']]
success_flag = True
remark = ""


def create_server_connection(host_name, user_name, password, db_name):
    connect = None
    try:
        connect = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=password,
            database=db_name
        )
        print("Connection Successful")
    except Error as err:
        print(f"Error: {err}")
    return connect


def execute_query(connect, query):
    cursor = connect.cursor()
    try:
        cursor.execute(query)
        connect.commit()
        print("Query Execute Successfully")
    except Error as err:
        print(f"Error : {err}")


def read_query(connect, query):
    cursor = connect.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()
    except Error as err:
        print(f"Error: {err}")
    return result


def check_email(s):
    pat = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    if re.match(pat, s):
        return True
    else:
        return False


CNADIDATE_SKILL_DICT = {}


def file_validation(row, column):
    for j in range(2, row + 1):
        success_flag = True
        # time.sleep(2)
        row_list = []
        remark = ""
        skill_list = []
        for i in range(1, column + 1):
            # print(sheet_obj.cell(row=j, column=i).value)
            if sheet_obj.cell(row=1, column=i).value == 'name':
                # print('name: ', type(sheet_obj.cell(row=j, column=i).value))
                name = sheet_obj.cell(row=j, column=i).value
                row_list.append(name)
                if name is None or type(name) != str:
                    success_flag = False
                    remark += "Invalid name, "
            if sheet_obj.cell(row=1, column=i).value == 'mob no':
                # print('mob no: ', type(sheet_obj.cell(row=j, column=i).value))
                mob_no = sheet_obj.cell(row=j, column=i).value
                row_list.append(mob_no)
                if mob_no is None or type(mob_no) != int or len(str(mob_no)) < 10:
                    success_flag = False
                    remark += "Invalid Mob No., "
            if sheet_obj.cell(row=1, column=i).value == 'email':
                # print('email: ', type(sheet_obj.cell(row=j, column=i).value))
                email_id = sheet_obj.cell(row=j, column=i).value
                row_list.append(email_id)
                if email_id is None or check_email(email_id) is False:
                    success_flag = False
                    remark += "Invalid Email-ID, "
            if sheet_obj.cell(row=1, column=i).value == 'created time':
                # print('created time: ', type(sheet_obj.cell(row=j, column=i).value))
                created_time = sheet_obj.cell(row=j, column=i).value
                row_list.append(created_time)
                if created_time is None or type(created_time) != datetime:
                    success_flag = False
                    remark += "Invalid Created Time, "
            if sheet_obj.cell(row=1, column=i).value == 'testscore':
                # print('testscore: ', type(sheet_obj.cell(row=j, column=i).value))
                test_score = sheet_obj.cell(row=j, column=i).value
                row_list.append(test_score)
                if test_score is None or (type(test_score) != float and type(test_score) != int):
                    remark += "Invalid Test Score, "
            if sheet_obj.cell(row=1, column=i).value == 'linked resume':
                # print('linked resume: ', type(sheet_obj.cell(row=j, column=i).value))
                linked_resume = sheet_obj.cell(row=j, column=i).value
                row_list.append(linked_resume)
                if linked_resume is None or type(linked_resume) != str:
                    remark += "Invalid Linked Resume, "
            if sheet_obj.cell(row=1, column=i).value == 'notice period':
                # print('notice period: ', type(sheet_obj.cell(row=j, column=i).value))
                notice_period = sheet_obj.cell(row=j, column=i).value
                row_list.append(notice_period)
                if notice_period is None or (type(notice_period) != float and type(notice_period) != int):
                    remark += "Invalid Notice Period, "
            if sheet_obj.cell(row=1, column=i).value == 'current ctc':
                # print('current ctc: ', type(sheet_obj.cell(row=j, column=i).value))
                current_ctc = sheet_obj.cell(row=j, column=i).value
                row_list.append(current_ctc)
                if current_ctc is None or (type(current_ctc) != float and type(current_ctc) != int):
                    remark += "Invalid Current CTC, "
            if sheet_obj.cell(row=1, column=i).value == 'expected ctc':
                # print('expected ctc: ', type(sheet_obj.cell(row=j, column=i).value))
                expected_ctc = sheet_obj.cell(row=j, column=i).value
                row_list.append(expected_ctc)
                if expected_ctc is None or (type(expected_ctc) != float and type(expected_ctc) != int):
                    remark += "Invalid Expected CTC, "
            if sheet_obj.cell(row=1, column=i).value == 'skill1':
                # print('skill1: ', type(sheet_obj.cell(row=j, column=i).value))
                skill1 = sheet_obj.cell(row=j, column=i).value
                row_list.append(skill1)
                if skill1 is None or type(skill1) != str:
                    remark += "Invalid Skill1, "
                    row_list.append(None)
                else:
                    query = f""" SELECT Id FROM skills WHERE Title = '{skill1}'"""
                    skill_result = read_query(connection, query)
                    if len(skill_result) > 0:
                        # print(skill_result)
                        skill_list.append({str(skill1).upper(): int(skill_result[0][0])})
                        row_list.append(int(skill_result[0][0]))
                        # CNADIDATE_SKILL_DICT[mob_no] = [skill1, skill_result]
                    else:
                        row_list.append(None)
                        print("SKILL1 NOT FOUND")
                        success_flag = False
                        remark += "skill1 is not present, "
            if sheet_obj.cell(row=1, column=i).value == 'skill1 years':
                # print('skill1 years: ', type(sheet_obj.cell(row=j, column=i).value))
                skill1_years = sheet_obj.cell(row=j, column=i).value
                row_list.append(skill1_years)
                if skill1_years is None or (type(skill1_years) != float and type(skill1_years) != int):
                    remark += "Invalid Skill1 Years, "
            if sheet_obj.cell(row=1, column=i).value == 'skill2':
                # print('skill2: ', type(sheet_obj.cell(row=j, column=i).value))
                skill2 = sheet_obj.cell(row=j, column=i).value
                row_list.append(skill2)
                if skill2 is None or type(skill2) != str:
                    remark += "Invalid Skill2, "
                    row_list.append(None)
                else:
                    query = f""" SELECT Id FROM skills WHERE Title = '{skill2}'"""
                    skill_result = read_query(connection, query)
                    if len(skill_result) > 0:
                        # print(skill_result)
                        skill_list.append({str(skill2).upper(): int(skill_result[0][0])})
                        row_list.append(int(skill_result[0][0]))
                        # CNADIDATE_SKILL_DICT[mob_no]= [skill2, skill_result]
                    else:
                        row_list.append(None)
                        print("SKILL2 NOT FOUND")
                        success_flag = False
                        remark += "skill2 is not present, "
            if sheet_obj.cell(row=1, column=i).value == 'skill2 years':
                # print('skill2 years: ', type(sheet_obj.cell(row=j, column=i).value))
                skill2_years = sheet_obj.cell(row=j, column=i).value
                row_list.append(skill2_years)
                if skill2_years is None or (type(skill2_years) != float and type(skill2_years) != int):
                    remark += "Invalid Skill2 Years, "
            if sheet_obj.cell(row=1, column=i).value == 'skill 3':
                # print('skill 3: ', type(sheet_obj.cell(row=j, column=i).value))
                skill3 = sheet_obj.cell(row=j, column=i).value
                row_list.append(skill3)
                if skill3 is None or type(skill3) != str:
                    remark += "Invalid Skill3, "
                    row_list.append(None)
                else:
                    query = f""" SELECT Id FROM skills WHERE Title = '{skill3}'"""
                    skill_result = read_query(connection, query)
                    if len(skill_result) > 0:
                        # print(skill_result)
                        skill_list.append({str(skill3).upper(): int(skill_result[0][0])})
                        row_list.append(int(skill_result[0][0]))
                        # CNADIDATE_SKILL_DICT[mob_no] = [skill3, skill_result]
                    else:
                        row_list.append(None)
                        print("SKILL3 NOT FOUND")
                        success_flag = False
                        remark += "skill3 is not present, "
            if sheet_obj.cell(row=1, column=i).value == 'skill3 years':
                # print('skill3 years: ', type(sheet_obj.cell(row=j, column=i).value))
                skill3_years = sheet_obj.cell(row=j, column=i).value
                row_list.append(skill3_years)
                if skill3_years is None or (type(skill3_years) != float and type(skill3_years) != int):
                    remark += "Invalid Skill3 Years, "
        print(skill_list)
        if len(skill_list) > 0:
            count = 0
            for skill in skill_list:
                if skill in job_skill_required:
                    count += 1
            if count > 0:
                print(count)
            else:
                success_flag = False
                remark += "Given Skills does not match with required job skills"
        row_list.append(remark.removesuffix(', '))
        print(row_list)
        print('-------------------------------------------------------------------------------------------------------')
        if success_flag:
            print(success_flag)
            print(row_list)
            success_row_list.append(row_list)
        else:
            print(success_flag)
            print(row_list)
            error_row_list.append(row_list)


# Give the location of the file
path = ""

pw = "Smprajkta4$"

db = "jobztop"

connection = create_server_connection("localhost", "root", pw, db)

q1 = '''
select * From upload;
'''

results = read_query(connection, q1)

for result in results:
    print(result)

q2 = '''
select user_id, path, job_id From upload where processed_state = 0;
'''

results = read_query(connection, q2)

if results is not None:
    for result in results:
        print(result)
        print(type(result[1]))
        user_id = result[0]
        path = str(result[1])
        job_id = int(result[2])
else:
    print("Results is None")

q2 = f'''
select SkillName, SkillId From jobskill where JobId = {job_id};
'''

job_skill_result = read_query(connection, q2)


print(f"job_skill_required: {job_skill_result}")

job_skill_required = []

if job_skill_result is not None:
    for result in job_skill_result:
        job_skill_required.append({str(result[0]): int(result[1])})
else:
    print("Results is None")

print(f"job_skill_required: {job_skill_required}")

print(path)
path = path
print(path)
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows:", row)
print("Total Columns:", column)

file_validation(row, column)

print(success_row_list)
print(error_row_list)


success_file_name = path.removesuffix('.xlsx') + '-Success.xlsx'
error_file_name = path.removesuffix('.xlsx') + '-Error.xlsx'
# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
workbook = Workbook()

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.

workbook.save(filename=success_file_name)

wb_success = openpyxl.load_workbook(success_file_name)

sheet_success = wb_success.active

for rows in tuple(success_row_list):
    sheet_success.append(tuple(rows))

wb_success.save(success_file_name)

workbook = Workbook()
workbook.save(filename=error_file_name)

wb_error = openpyxl.load_workbook(error_file_name)

sheet_error = wb_error.active

for rows in tuple(error_row_list):
    sheet_error.append(tuple(rows))

wb_error.save(error_file_name)


q3 = f'''
UPDATE upload
SET processed_state = 1
WHERE user_id = '{user_id}';
'''

execute_query(connection, q3)

time_now = datetime.now()
time_now = time_now.strftime('%Y-%m-%d %H:%M:%S')
query = f'''
        UPDATE upload
        SET file_update_time = '{time_now}'
        WHERE user_id = '{user_id}';
        '''
execute_query(connection, query)

print(CNADIDATE_SKILL_DICT)

q1 = '''
select * From upload;
'''

'''results = read_query(connection, q1)

for result in results:
    print(result)'''
